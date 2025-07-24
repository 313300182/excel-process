# -*- coding: utf-8 -*-
"""
工资Excel读取器
从源Excel文件中提取工资相关数据
"""

import logging
import re
from typing import Dict, Any, Optional, List
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

from config.salary_settings import SALARY_CONFIG


class SalaryExcelReader:
    """工资Excel读取器"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.config = SALARY_CONFIG['source_extraction']
        
    def read_salary_data(self, file_path: str) -> Dict[str, Any]:
        """
        读取工资数据文件（从业务分组处理后的文件中读取）
        
        Args:
            file_path: Excel文件路径（业务分组处理后的文件）
            
        Returns:
            Dict[str, Any]: 提取的员工工资数据
        """
        workbook = None
        try:
            self.logger.info(f"🔧 准备加载Excel文件: {file_path}")
            # 使用read_only模式减少内存使用
            self.logger.debug("📖 开始加载workbook...")
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
            self.logger.info(f"✅ Excel文件加载成功")
            all_employees_data = []
            
            # 先获取sheet列表，避免在迭代中保持所有sheet引用
            self.logger.debug("📋 获取sheet列表...")
            sheet_names = list(workbook.sheetnames)
            self.logger.info(f"📊 找到 {len(sheet_names)} 个sheet: {sheet_names}")
            
            # 遍历所有sheet，每个sheet代表一个员工
            self.logger.info("🔄 开始遍历sheet...")
            for i, sheet_name in enumerate(sheet_names):
                try:
                    self.logger.info(f"📄 处理第 {i+1}/{len(sheet_names)} 个sheet: {sheet_name}")
                    worksheet = workbook[sheet_name]
                    self.logger.info(f"✅ 成功访问worksheet: {sheet_name}")
                    
                    # 从sheet名称提取员工姓名（去掉角色标识）
                    self.logger.info(f"🔍 提取员工姓名从: {sheet_name}")
                    employee_name = self._extract_employee_name_from_sheet(sheet_name)
                    if not employee_name:
                        self.logger.info(f"⏭️ 跳过sheet: {sheet_name} (非员工sheet)")
                        continue
                        
                    self.logger.info(f"👤 处理员工: {employee_name} (Sheet: {sheet_name})")
                    
                    # 提取业绩数据
                    self.logger.info(f"📊 开始提取业绩数据...")
                    performance_data = self._extract_performance_data(worksheet)
                    self.logger.info(f"✅ 业绩数据提取完成: {type(performance_data)}")
                    
                    if performance_data:
                        employee_data = {
                            'employee_info': {
                                'name': employee_name,
                                'sheet_name': sheet_name,
                                'month': self._extract_month_info(worksheet)
                            },
                            'performance_data': performance_data,
                            'file_path': file_path
                        }
                        all_employees_data.append(employee_data)
                        
                        # 打印到日志
                        total_performance = performance_data.get('total_performance_value', 0)
                        self.logger.info(f"员工 {employee_name}: 实收业绩+体验卡合计 = {total_performance}")
                    
                    # 清理worksheet引用
                    worksheet = None
                    
                except Exception as e:
                    self.logger.warning(f"处理Sheet {sheet_name} 失败: {str(e)}")
                    continue
            
            self.logger.info(f"成功读取 {len(all_employees_data)} 个员工的工资数据")
            
            return {
                'employees': all_employees_data,
                'total_count': len(all_employees_data),
                'file_path': file_path
            }
            
        except Exception as e:
            self.logger.error(f"读取文件失败 {file_path}: {str(e)}")
            raise Exception(f"读取文件失败: {str(e)}")
        finally:
            # 确保workbook被正确关闭，释放内存
            if workbook:
                try:
                    workbook.close()
                    workbook = None
                except:
                    pass
            
    def _extract_employee_name_from_sheet(self, sheet_name: str) -> str:
        """
        从sheet名称中提取员工姓名
        去掉角色标识，如 "张三(服务老师)" -> "张三"
        
        Args:
            sheet_name: sheet名称
            
        Returns:
            str: 员工姓名
        """
        try:
            # 去掉角色标识
            if '(' in sheet_name and ')' in sheet_name:
                name = sheet_name.split('(')[0].strip()
            else:
                name = sheet_name.strip()
                
            # 排除非员工的sheet（如未分类等）
            excluded_names = [
                '未分类', '汇总', '统计', 'Sheet1', 'Sheet', 'Sheet2', 'Sheet3',
                '合计', '总计', '小计', '数据', '备注', '说明', '模板', 'Template',
                '测试', 'Test', '总览', '概要', '摘要', '目录', '索引'
            ]
            
            if name in excluded_names or not name.strip():
                self.logger.info(f"跳过非员工Sheet: {sheet_name}")
                return None
                
            return name if name else None
            
        except Exception as e:
            self.logger.warning(f"提取员工姓名失败 {sheet_name}: {str(e)}")
            return None
            
    def _extract_month_info(self, worksheet) -> str:
        """
        从工作表中提取月份信息
        
        Args:
            worksheet: Excel工作表
            
        Returns:
            str: 月份信息
        """
        try:
            # 尝试从第一行查找包含月份的信息
            for row in worksheet.iter_rows(min_row=1, max_row=3, values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str):
                        if '年' in cell and '月' in cell:
                            return cell.strip()
            
            # 如果没找到，返回默认值
            import datetime
            return datetime.datetime.now().strftime("%Y年%m月")
            
        except Exception as e:
            self.logger.warning(f"提取月份信息失败: {str(e)}")
            return "未知月份"
            
    def _extract_performance_data(self, worksheet) -> Dict[str, Any]:
        """
        从工作表中提取业绩数据
        直接从L列找最后一个非空值作为总业绩
        
        Args:
            worksheet: Excel工作表
            
        Returns:
            Dict[str, Any]: 业绩数据
        """
        try:
            self.logger.info("🧮 开始提取业绩数据")
            performance_data = {
                'actual_revenue': 0,
                'experience_card': 0,
                'total_performance_value': 0
            }
            
            # 直接从L列(第12列)找最后一个非空值
            l_column = 12  # L列是第12列
            last_value = 0
            
            self.logger.info(f"📊 从L列扫描数据，工作表最大行数: {worksheet.max_row}")
            
            # 从底部往上找最后一个非空值
            scan_count = 0
            for row_idx in range(worksheet.max_row, 0, -1):
                scan_count += 1
                if scan_count % 50 == 0:  # 每50行记录一次
                    self.logger.info(f"🔍 已扫描 {scan_count} 行，当前行: {row_idx}")
                
                try:
                    cell_value = worksheet.cell(row=row_idx, column=l_column).value
                    if cell_value is not None:
                        last_value = self._convert_to_number(cell_value)
                        if last_value != 0:
                            self.logger.info(f"✅ 从L{row_idx}找到最后的值: {last_value}")
                            break
                except Exception as e:
                    self.logger.warning(f"访问L{row_idx}单元格失败: {str(e)}")
                    continue
            
            self.logger.info(f"📈 扫描完成，共扫描 {scan_count} 行")
            
            # 将找到的值作为总业绩
            performance_data['actual_revenue'] = last_value
            performance_data['experience_card'] = 0
            
            # 计算总和并乘以10000
            total_value = last_value * 10000
            performance_data['total_performance_value'] = total_value
            
            self.logger.info(f"💰 L列最后值: {last_value}, 计算值: {total_value}")
            
            return performance_data
            
        except Exception as e:
            self.logger.error(f"提取业绩数据失败: {str(e)}")
            import traceback
            self.logger.error(f"异常详情:\n{traceback.format_exc()}")
            return {'actual_revenue': 0, 'experience_card': 0, 'total_performance_value': 0}
            
    def _convert_to_number(self, value) -> float:
        """
        将值转换为数字
        
        Args:
            value: 要转换的值
            
        Returns:
            float: 转换后的数字
        """
        if value is None:
            return 0.0
            
        try:
            if isinstance(value, (int, float)):
                return float(value)
            elif isinstance(value, str):
                # 移除逗号和空格
                cleaned = str(value).replace(',', '').replace(' ', '').strip()
                if cleaned:
                    return float(cleaned)
            return 0.0
        except (ValueError, TypeError):
            return 0.0
            
    def _extract_employee_info(self, worksheet) -> Dict[str, Any]:
        """
        提取员工基本信息
        
        Args:
            worksheet: Excel工作表
            
        Returns:
            Dict[str, Any]: 员工信息
        """
        employee_info = {}
        
        try:
            # 提取姓名
            name_cell = self.config['employee_info']['name_cell']
            name_value = worksheet[name_cell].value
            if name_value:
                employee_info['name'] = str(name_value).strip()
            
            # 提取月份
            month_cell = self.config['employee_info']['month_cell']
            month_value = worksheet[month_cell].value
            if month_value:
                employee_info['month'] = str(month_value).strip()
            
            self.logger.info(f"提取员工信息: {employee_info}")
            
        except Exception as e:
            self.logger.warning(f"提取员工信息时出错: {str(e)}")
            
        return employee_info
        
    def _extract_salary_details(self, worksheet) -> List[Dict[str, Any]]:
        """
        提取工资明细数据
        
        Args:
            worksheet: Excel工作表
            
        Returns:
            List[Dict[str, Any]]: 工资明细列表
        """
        details = []
        
        start_row = self.config['data_start_row']
        start_col = self.config['data_start_col']
        max_rows = self.config['max_rows']
        end_markers = self.config['end_markers']
        fields = self.config['fields']
        
        try:
            for row_idx in range(start_row, start_row + max_rows):
                # 检查是否到达结束标记
                if self._is_end_marker(worksheet, row_idx, start_col, end_markers):
                    break
                
                # 提取当前行数据
                row_data = {}
                is_empty_row = True
                
                for field_name, col_offset in fields.items():
                    col_idx = start_col + col_offset - 1
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                    
                    if cell_value is not None:
                        processed_value = self._process_cell_value(cell_value, field_name)
                        row_data[field_name] = processed_value
                        is_empty_row = False
                    else:
                        row_data[field_name] = None
                
                # 跳过空行
                if not is_empty_row and self._is_valid_row(row_data):
                    details.append(row_data)
                    
        except Exception as e:
            self.logger.error(f"提取工资明细时出错: {str(e)}")
            
        return details
        
    def _process_cell_value(self, value: Any, field_name: str) -> Any:
        """
        处理单元格值
        
        Args:
            value: 原始值
            field_name: 字段名
            
        Returns:
            Any: 处理后的值
        """
        if value is None:
            return None
            
        # 转换为字符串处理
        str_value = str(value).strip()
        
        # 数字类型字段处理
        if field_name in ['quantity', 'rate', 'amount']:
            return self._process_number_value(str_value)
        
        # 字符串类型字段
        return str_value
        
    def _process_number_value(self, value_str: str) -> float:
        """
        处理数字值
        
        Args:
            value_str: 字符串值
            
        Returns:
            float: 数字值
        """
        if not value_str or value_str == '':
            return 0.0
            
        try:
            # 移除逗号分隔符
            cleaned_value = value_str.replace(',', '')
            
            # 尝试转换为数字
            if '.' in cleaned_value:
                return float(cleaned_value)
            else:
                return float(int(cleaned_value))
                
        except (ValueError, TypeError):
            self.logger.warning(f"无法转换数字值: {value_str}")
            return 0.0
            
    def _is_end_marker(self, worksheet, row_idx: int, start_col: int, 
                      end_markers: List[str]) -> bool:
        """
        检查是否为结束标记
        
        Args:
            worksheet: 工作表
            row_idx: 行索引
            start_col: 开始列
            end_markers: 结束标记列表
            
        Returns:
            bool: 是否为结束标记
        """
        try:
            # 检查前几列是否包含结束标记
            for col_offset in range(5):
                col_idx = start_col + col_offset
                cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                
                if cell_value:
                    str_value = str(cell_value).strip()
                    for marker in end_markers:
                        if marker in str_value:
                            return True
                            
        except Exception as e:
            self.logger.debug(f"检查结束标记时出错: {str(e)}")
            
        return False
        
    def _is_valid_row(self, row_data: Dict[str, Any]) -> bool:
        """
        检查行数据是否有效
        
        Args:
            row_data: 行数据
            
        Returns:
            bool: 是否有效
        """
        # 至少项目字段不为空
        project = row_data.get('project')
        return project is not None and str(project).strip() != ''
        
    def _calculate_statistics(self, details: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        计算统计数据
        
        Args:
            details: 工资明细
            
        Returns:
            Dict[str, Any]: 统计数据
        """
        statistics = {
            'total_records': len(details),
            'categories': {},
            'total_amounts': {}
        }
        
        key_fields = self.config['key_fields']
        
        try:
            # 统计各类别数据
            for detail in details:
                category = detail.get('category', '未知')
                project = detail.get('project', '')
                amount = detail.get('amount', 0) or 0
                
                # 统计类别
                if category not in statistics['categories']:
                    statistics['categories'][category] = {'count': 0, 'total': 0}
                statistics['categories'][category]['count'] += 1
                statistics['categories'][category]['total'] += float(amount)
                
                # 识别关键项目
                for key, keyword in key_fields.items():
                    if keyword in str(project):
                        statistics['total_amounts'][key] = float(amount)
                        break
                        
        except Exception as e:
            self.logger.error(f"计算统计数据时出错: {str(e)}")
            
        return statistics
        
    def validate_file_structure(self, file_path: str) -> bool:
        """
        验证文件结构是否符合要求（业务分组处理后的文件）
        
        Args:
            file_path: 文件路径
            
        Returns:
            bool: 是否符合要求
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # 检查是否至少有一个有效的员工sheet
            valid_employee_count = 0
            
            for sheet_name in workbook.sheetnames:
                employee_name = self._extract_employee_name_from_sheet(sheet_name)
                if employee_name:
                    worksheet = workbook[sheet_name]
                    
                    # 检查是否有实收业绩和体验卡列
                    has_performance_data = self._check_performance_columns(worksheet)
                    if has_performance_data:
                        valid_employee_count += 1
                        
            workbook.close()
            return valid_employee_count > 0
            
        except Exception as e:
            self.logger.error(f"验证文件结构时出错 {file_path}: {str(e)}")
            return False
            
    def _check_performance_columns(self, worksheet) -> bool:
        """
        检查工作表是否包含实收业绩和体验卡列
        
        Args:
            worksheet: Excel工作表
            
        Returns:
            bool: 是否包含必要的列
        """
        try:
            has_actual_revenue = False
            has_experience_card = False
            
            # 在前几行中查找表头
            for row in worksheet.iter_rows(min_row=1, max_row=5, values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str):
                        cell_str = str(cell).strip()
                        if '实收业绩' in cell_str:
                            has_actual_revenue = True
                        elif '体验卡' in cell_str:
                            has_experience_card = True
                            
            return has_actual_revenue or has_experience_card  # 至少有一个即可
            
        except Exception as e:
            self.logger.warning(f"检查业绩列时出错: {str(e)}")
            return False 