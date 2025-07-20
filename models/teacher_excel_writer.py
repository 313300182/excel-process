# -*- coding: utf-8 -*-
"""
老师分组Excel写入器模型
负责将老师分组数据写入多个sheet的Excel文件
"""

import os
import logging
import shutil
import stat
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from config.teacher_splitter_settings import (
    TEACHER_OUTPUT_CONFIG, 
    TEACHER_FILE_CONFIG
)


class TeacherExcelWriter:
    """老师分组Excel文件写入器"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
    def create_teacher_grouped_file(self, 
                                   all_data: List[Dict[str, Any]], 
                                   output_dir: str, 
                                   original_filename: str,
                                   source_file_path: str) -> Optional[str]:
        """
        创建按老师角色分类的多个Excel文件
        
        Args:
            all_data: 所有数据列表
            output_dir: 输出目录
            original_filename: 原始文件名
            source_file_path: 源文件路径
            
        Returns:
            str: 输出目录路径，失败返回None
        """
        try:
            if not all_data:
                self.logger.warning("没有数据需要写入")
                return None
                
            # 确保输出目录存在
            os.makedirs(output_dir, exist_ok=True)
            
            # 按角色和店家分组数据
            grouped_data = self._group_data_by_role(all_data)
            
            # 按角色类型重新分组
            role_files = {
                '服务总监': {},
                '服务老师': {},
                '操作老师': {},
                '店家': {}
            }
            
            # 将数据按角色类型分类
            for group_key, role_data in grouped_data.items():
                if '(服务总监)' in group_key:
                    role_files['服务总监'][group_key] = role_data
                elif '(服务老师)' in group_key:
                    role_files['服务老师'][group_key] = role_data
                elif '(操作老师)' in group_key:
                    role_files['操作老师'][group_key] = role_data
                elif '(店家)' in group_key:
                    role_files['店家'][group_key] = role_data
            
            created_files = []
            
            # 为每个角色类型创建独立的Excel文件
            for role_type, role_groups in role_files.items():
                if not role_groups:  # 跳过空数据的角色
                    continue
                
                # 生成该角色类型的文件名
                role_filename = self._generate_role_filename(original_filename, role_type)
                role_output_path = os.path.join(output_dir, role_filename)
                
                # 创建该角色类型的工作簿
                workbook = Workbook()
                
                # 删除默认的sheet
                if workbook.worksheets:
                    workbook.remove(workbook.worksheets[0])
                
                sheet_count = 0
                
                # 为该角色类型下的每个具体老师/店家创建sheet
                for group_key, group_data in role_groups.items():
                    if not group_data:
                        continue
                    
                    # 提取sheet名称（去掉角色标识）
                    sheet_name = group_key.split('(')[0]
                    worksheet = workbook.create_sheet(sheet_name)
                    
                    # 写入表头
                    self._write_headers(worksheet)
                    
                    # 写入数据
                    self._write_teacher_data(worksheet, group_data)
                    
                    # 添加合计行
                    if TEACHER_OUTPUT_CONFIG.get('add_total_row', False):
                        self._write_total_row(worksheet, group_data)
                    
                    # 设置列宽
                    self._adjust_column_width(worksheet)
                    
                    sheet_count += 1
                    self.logger.info(f"在{role_type}文件中创建sheet: {sheet_name}, 数据行数: {len(group_data)}")
                
                # 保存该角色类型的文件
                workbook.save(role_output_path)
                workbook.close()
                created_files.append(role_output_path)
                
                self.logger.info(f"成功创建{role_type}文件: {role_output_path}, 共{sheet_count}个sheet")
            
            self.logger.info(f"成功创建所有老师角色分类文件，共{len(created_files)}个文件")
            return output_dir  # 返回输出目录
            
        except Exception as e:
            self.logger.error(f"创建老师角色分类文件失败: {e}")
            return None

    
    def _group_data_by_role(self, all_data: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
        """
        按角色和店家分组数据，为每个老师和店家创建独立分组（包括空值）
        
        Args:
            all_data: 所有数据
            
        Returns:
            Dict: 分组后的数据，key为"老师姓名(角色)"或"店家名称(店家)"
        """
        grouped_data = {}
        
        role_mapping = {
            'service_director': '服务总监',
            'service_teacher': '服务老师',
            'operation_teacher': '操作老师'
        }
        
        empty_name = TEACHER_FILE_CONFIG['empty_teacher_name']
        
        for row in all_data:
            # 为每个角色创建具体老师的分组（包括空值）
            for role_field, role_name in role_mapping.items():
                teacher_name = row.get(role_field)
                
                # 处理有值和无值的情况
                if teacher_name and str(teacher_name).strip() != '':
                    teacher_name = str(teacher_name).strip()
                    group_key = f"{teacher_name}({role_name})"
                else:
                    # 空值数据归入"未分类"
                    group_key = f"未分类({role_name})"
                
                if group_key not in grouped_data:
                    grouped_data[group_key] = []
                
                grouped_data[group_key].append(row)
            
            # 按店家分组（包括空值）
            store_name = row.get('store_name')
            if store_name and str(store_name).strip() != '':
                store_name = str(store_name).strip()
                group_key = f"{store_name}(店家)"
            else:
                # 空店家名称归入"未分类"
                group_key = f"未分类(店家)"
            
            if group_key not in grouped_data:
                grouped_data[group_key] = []
            
            grouped_data[group_key].append(row)
        
        # 按分组key排序（按角色和姓名排序）
        sorted_groups = dict(sorted(grouped_data.items()))
        
        return sorted_groups
    

    
    def _generate_output_filename(self, original_filename: str) -> str:
        """
        生成输出文件名
        
        Args:
            original_filename: 原始文件名
            
        Returns:
            str: 生成的文件名
        """
        name_without_ext = os.path.splitext(original_filename)[0]
        timestamp = datetime.now().strftime(TEACHER_FILE_CONFIG['timestamp_format'])
        
        filename = TEACHER_FILE_CONFIG['filename_format'].format(
            original_name=name_without_ext,
            timestamp=timestamp
        )
        
        return filename

    def _generate_role_filename(self, original_filename: str, role_type: str) -> str:
        """
        生成按角色类型分类的文件名
        
        Args:
            original_filename: 原始文件名
            role_type: 角色类型 (如 "服务总监", "服务老师", "操作老师", "店家")
            
        Returns:
            str: 生成的文件名
        """
        name_without_ext = os.path.splitext(original_filename)[0]
        timestamp = datetime.now().strftime(TEACHER_FILE_CONFIG['timestamp_format'])
        
        filename = TEACHER_FILE_CONFIG['filename_format'].format(
            original_name=f"{name_without_ext}_{role_type}",
            timestamp=timestamp
        )
        
        return filename
    
    def _write_headers(self, worksheet) -> None:
        """
        写入表头
        
        Args:
            worksheet: 工作表对象
        """
        try:
            headers = TEACHER_OUTPUT_CONFIG['headers']
            
            # 创建表头样式 - 增加字体大小和边框
            header_font = Font(bold=True, color="FFFFFF", size=14)  # 增加字体大小
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 添加边框样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for (row, col), header_text in headers.items():
                cell = worksheet.cell(row=row, column=col, value=header_text)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                
            # 设置表头行高
            worksheet.row_dimensions[1].height = 25  # 增加行高以配合更大字体
                
        except Exception as e:
            self.logger.error(f"写入表头失败: {e}")
    
    def _write_teacher_data(self, worksheet, teacher_data: List[Dict[str, Any]]) -> int:
        """
        写入老师数据
        
        Args:
            worksheet: 工作表对象
            teacher_data: 老师数据列表
            
        Returns:
            int: 最后写入的行号
        """
        try:
            output_columns = TEACHER_OUTPUT_CONFIG['output_columns']
            data_start_row = TEACHER_OUTPUT_CONFIG['data_start_row']
            
            # 数值字段列表
            numeric_fields = ['order_amount', 'debt_collection', 'payment', 'card_deduction', 
                            'debt', 'commission', 'experience_card', 'public_revenue']
            
            # 创建数据行样式
            data_font = Font(size=12)  # 增加数据字体大小
            data_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row_index, data_row in enumerate(teacher_data):
                current_row = data_start_row + row_index
                
                # 设置数据行高
                worksheet.row_dimensions[current_row].height = 20
                
                for field_name, col_index in output_columns.items():
                    value = data_row.get(field_name, "")
                    
                    # 处理数值字段
                    if field_name in numeric_fields and value:
                        try:
                            if isinstance(value, str):
                                value = value.replace(',', '').replace('，', '')
                            value = float(value) if value else 0
                        except (ValueError, TypeError):
                            value = 0
                    
                    cell = worksheet.cell(row=current_row, column=col_index, value=value)
                    
                    # 应用样式 - 所有内容都居中显示
                    cell.font = data_font
                    cell.alignment = data_alignment  # 统一居中对齐
                    cell.border = thin_border
                    
            return data_start_row + len(teacher_data) - 1
                        
        except Exception as e:
            self.logger.error(f"写入老师数据失败: {e}")
            return data_start_row
    
    def _write_total_row(self, worksheet, teacher_data: List[Dict[str, Any]]) -> None:
        """
        写入合计行
        
        Args:
            worksheet: 工作表对象
            teacher_data: 老师数据列表
        """
        try:
            data_start_row = TEACHER_OUTPUT_CONFIG['data_start_row']
            total_row = data_start_row + len(teacher_data)
            
            total_label = TEACHER_OUTPUT_CONFIG['total_label']
            total_label_column = TEACHER_OUTPUT_CONFIG['total_label_column']
            total_columns = TEACHER_OUTPUT_CONFIG['total_columns']
            
            # 写入合计标签
            worksheet.cell(row=total_row, column=total_label_column, value=total_label)
            
            # 计算各列合计金额
            totals = {}
            for field_name in total_columns.keys():
                totals[field_name] = 0
                
                for data_row in teacher_data:
                    value = data_row.get(field_name, 0)
                    if value:
                        try:
                            if isinstance(value, str):
                                value = value.replace(',', '').replace('，', '')
                            totals[field_name] += float(value)
                        except (ValueError, TypeError):
                            pass
            
            # 写入各列合计数据
            for field_name, col_index in total_columns.items():
                worksheet.cell(row=total_row, column=col_index, value=totals[field_name])
            
            # 设置合计行样式 - 美化
            total_font = Font(bold=True, size=12)  # 增加字体大小
            total_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            total_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 设置合计行高
            worksheet.row_dimensions[total_row].height = 22
            
            for col in range(1, 16):  # A-O列
                cell = worksheet.cell(row=total_row, column=col)
                cell.font = total_font
                cell.fill = total_fill
                cell.alignment = total_alignment  # 统一居中对齐
                cell.border = thin_border
            
            # 添加实收业绩和体验卡合计行
            grand_total_row = total_row + 1
            grand_total = totals['commission'] + totals['experience_card']
            
            # 写入总合计标签和数值
            worksheet.cell(row=grand_total_row, column=total_label_column, value="实收业绩和体验卡合计")
            worksheet.cell(row=grand_total_row, column=12, value=grand_total)  # 在实收业绩列显示总合计
            
            # 设置总合计行样式 - 美化
            grand_total_font = Font(bold=True, color="FF0000", size=12)  # 红色字体，增加字体大小
            grand_total_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # 黄色背景
            grand_total_alignment = Alignment(horizontal="center", vertical="center")
            
            # 设置总合计行高
            worksheet.row_dimensions[grand_total_row].height = 22
            
            for col in range(1, 16):  # A-O列
                cell = worksheet.cell(row=grand_total_row, column=col)
                cell.font = grand_total_font
                cell.fill = grand_total_fill
                cell.alignment = grand_total_alignment  # 统一居中对齐
                cell.border = thin_border
            
        except Exception as e:
            self.logger.error(f"写入合计行失败: {e}")
    
    def _adjust_column_width(self, worksheet) -> None:
        """
        调整列宽
        
        Args:
            worksheet: 工作表对象
        """
        try:
            # 设置各列宽度 - 适配更多列和更大字体
            column_widths = {
                'A': 12,  # 日期
                'B': 15,  # 客户
                'C': 12,  # 服务总监
                'D': 12,  # 服务老师
                'E': 12,  # 操作老师
                'F': 20,  # 店名
                'G': 12,  # 开单金额
                'H': 12,  # 收欠款
                'I': 12,  # 收款
                'J': 12,  # 卡扣
                'K': 12,  # 欠款
                'L': 15,  # 实收业绩
                'M': 12,  # 体验卡
                'N': 42,  # 开单明细 - 增加宽度以容纳更多详细信息
                'O': 12,  # 公司收
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
                
        except Exception as e:
            self.logger.error(f"调整列宽失败: {e}") 