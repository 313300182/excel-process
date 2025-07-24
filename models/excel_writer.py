# -*- coding: utf-8 -*-
"""
Excel写入器模型
负责将数据填充到模板Excel文件中并生成新文件
"""

import os
import logging
import shutil
import stat
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.merge import MergedCellRange
from openpyxl.styles import Alignment

from config.settings import TARGET_CONFIG, OUTPUT_CONFIG, get_template_path


class ExcelWriter:
    """Excel文件写入器"""
    
    def __init__(self, template_path: Optional[str] = None):
        self.logger = logging.getLogger(__name__)
        self.template_path = template_path if template_path else get_template_path()
    
    def set_template_path(self, template_path: str):
        """
        设置模板文件路径
        
        Args:
            template_path: 模板文件路径
        """
        self.template_path = template_path
        
    def validate_template(self) -> bool:
        """
        验证模板文件是否存在和有效
        
        Returns:
            bool: 模板是否有效
        """
        try:
            if not os.path.exists(self.template_path):
                self.logger.error(f"模板文件不存在: {self.template_path}")
                return False
                
            # 尝试打开模板文件
            workbook = load_workbook(self.template_path)
            workbook.close()
            return True
            
        except InvalidFileException:
            self.logger.error(f"无效的模板文件: {self.template_path}")
            return False
        except Exception as e:
            self.logger.error(f"验证模板文件时发生错误: {e}")
            return False
    
    def _ensure_output_directory(self, output_dir: str) -> bool:
        """
        确保输出目录存在并且有写入权限
        
        Args:
            output_dir: 输出目录路径
            
        Returns:
            bool: 目录是否可用
        """
        try:
            # 创建目录（如果不存在）
            os.makedirs(output_dir, exist_ok=True)
            
            # 检查写入权限
            if not os.access(output_dir, os.W_OK):
                # 尝试修改权限
                try:
                    os.chmod(output_dir, stat.S_IWRITE | stat.S_IREAD)
                except PermissionError:
                    self.logger.error(f"无权限写入目录: {output_dir}")
                    return False
            
            # 测试写入权限
            test_file = os.path.join(output_dir, 'test_write.tmp')
            try:
                with open(test_file, 'w') as f:
                    f.write('test')
                os.remove(test_file)
                return True
            except PermissionError:
                self.logger.error(f"目录写入权限测试失败: {output_dir}")
                return False
                
        except Exception as e:
            self.logger.error(f"创建输出目录失败: {e}")
            return False
    
    def _safe_write_cell(self, worksheet, row: int, col: int, value: Any) -> bool:
        """
        安全地写入单元格（处理合并单元格）并设置居中对齐
        
        Args:
            worksheet: 工作表对象
            row: 行号
            col: 列号
            value: 要写入的值
            
        Returns:
            bool: 是否写入成功
        """
        try:
            cell = worksheet.cell(row=row, column=col)
            target_cell = cell  # 默认目标单元格
            
            # 检查是否是合并单元格
            if hasattr(cell, 'coordinate'):
                for merged_range in worksheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        # 这是合并单元格，获取左上角单元格
                        target_cell = worksheet.cell(
                            row=merged_range.min_row, 
                            column=merged_range.min_col
                        )
                        self.logger.debug(f"写入合并单元格 {cell.coordinate} -> {target_cell.coordinate}: {value}")
                        break
            
            # 写入值
            target_cell.value = value
            
            # 设置居中对齐
            target_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            self.logger.debug(f"写入单元格 {target_cell.coordinate}: {value} (居中对齐)")
            return True
            
        except Exception as e:
            self.logger.warning(f"写入单元格失败 ({row}, {col}): {e}")
            return False
    
    def create_output_file(self, data_list: List[Dict[str, Any]], 
                          output_dir: str, 
                          original_filename: str) -> Optional[str]:
        """
        创建输出文件（支持多行数据）
        
        Args:
            data_list: 要写入的数据列表
            output_dir: 输出目录
            original_filename: 原始文件名
            
        Returns:
            str: 生成的文件路径，失败返回None
        """
        try:
            if not self.validate_template():
                return None
                
            if not data_list:
                self.logger.warning("没有数据需要写入")
                return None
                
            # 确保输出目录存在并可写
            if not self._ensure_output_directory(output_dir):
                return None
            
            # 生成输出文件名
            output_filename = self._generate_output_filename(original_filename)
            output_path = os.path.join(output_dir, output_filename)
            
            # 检查文件是否已存在并且被占用
            if os.path.exists(output_path):
                try:
                    # 尝试删除已存在的文件
                    os.remove(output_path)
                except PermissionError:
                    # 如果无法删除，添加随机后缀
                    import time
                    name_parts = os.path.splitext(output_filename)
                    output_filename = f"{name_parts[0]}_{int(time.time())}{name_parts[1]}"
                    output_path = os.path.join(output_dir, output_filename)
            
            # 复制模板文件
            shutil.copy2(self.template_path, output_path)
            
            # 设置文件权限
            try:
                os.chmod(output_path, stat.S_IWRITE | stat.S_IREAD)
            except PermissionError:
                pass  # 忽略权限设置失败
            
            # 打开文件进行编辑
            workbook = load_workbook(output_path)
            worksheet = workbook.active
            
            # 写入表头（安全模式）
            self._write_headers_safe(worksheet)
            
            # 写入数据
            last_row = self._write_data_safe(worksheet, data_list)
            
            # 写入合计行（如果配置了）
            if TARGET_CONFIG.get('add_total_row', False):
                self._write_total_row_safe(worksheet, data_list, last_row + 1)
            
            # 保存文件
            workbook.save(output_path)
            workbook.close()
            
            self.logger.info(f"成功创建输出文件: {output_path}")
            return output_path
            
        except Exception as e:
            self.logger.error(f"创建输出文件失败: {e}")
            return None
    
    def create_multiple_files(self, data_list: List[Dict[str, Any]], 
                             output_dir: str, 
                             original_filename: str,
                             records_per_file: int = 100) -> List[str]:
        """
        创建多个输出文件（当数据量较大时）
        
        Args:
            data_list: 要写入的数据列表
            output_dir: 输出目录  
            original_filename: 原始文件名
            records_per_file: 每个文件的记录数
            
        Returns:
            List[str]: 生成的文件路径列表
        """
        try:
            if not data_list:
                return []
                
            output_files = []
            total_records = len(data_list)
            
            # 按批次处理数据
            for i in range(0, total_records, records_per_file):
                batch_data = data_list[i:i + records_per_file]
                batch_num = (i // records_per_file) + 1
                
                # 修改文件名以包含批次号
                name_parts = os.path.splitext(original_filename)
                batch_filename = f"{name_parts[0]}_batch_{batch_num}{name_parts[1]}"
                
                output_path = self.create_output_file(batch_data, output_dir, batch_filename)
                if output_path:
                    output_files.append(output_path)
                    
            return output_files
            
        except Exception as e:
            self.logger.error(f"创建多个输出文件失败: {e}")
            return []
    
    def _generate_output_filename(self, original_filename: str) -> str:
        """
        生成输出文件名
        
        Args:
            original_filename: 原始文件名
            
        Returns:
            str: 生成的文件名
        """
        # 移除扩展名
        name_without_ext = os.path.splitext(original_filename)[0]
        
        # 生成时间戳
        timestamp = datetime.now().strftime(OUTPUT_CONFIG['timestamp_format'])
        
        # 使用配置的格式生成文件名
        filename = OUTPUT_CONFIG['filename_format'].format(
            original_name=name_without_ext,
            timestamp=timestamp
        )
        
        return filename
    
    def _write_headers_safe(self, worksheet) -> None:
        """
        安全地写入表头（处理合并单元格）
        
        Args:
            worksheet: 工作表对象
        """
        try:
            headers = TARGET_CONFIG.get('headers', {})
            success_count = 0
            for (row, col), header_text in headers.items():
                if self._safe_write_cell(worksheet, row, col, header_text):
                    success_count += 1
                    
            self.logger.debug(f"写入表头: {success_count}/{len(headers)} 个成功")
                
        except Exception as e:
            self.logger.error(f"写入表头失败: {e}")
    
    def _write_data_safe(self, worksheet, data_list: List[Dict[str, Any]]) -> int:
        """
        安全地写入数据到工作表（支持多行数据，处理合并单元格）
        
        Args:
            worksheet: 工作表对象
            data_list: 数据列表
            
        Returns:
            int: 最后写入的行号
        """
        try:
            fill_positions = TARGET_CONFIG['fill_positions']
            data_start_row = TARGET_CONFIG.get('data_start_row', 2)
            
            # 获取默认值配置
            default_values = TARGET_CONFIG.get('default_values', {})
            
            for row_index, data_row in enumerate(data_list):
                current_row = data_start_row + row_index
                
                for field_name, col_index in fill_positions.items():
                    # 优先使用数据行中的值，如果没有则使用默认值
                    if field_name in data_row:
                        value = data_row[field_name]
                    elif field_name in default_values:
                        value = default_values[field_name]
                    else:
                        continue  # 跳过没有数据也没有默认值的字段
                        
                    # 处理空值
                    if value is None:
                        value = ""
                    # 处理数字格式（但排除字符串类型的字段）
                    elif field_name in ['quantity', 'amount', 'tax_rate'] and value != "":
                        try:
                            # 尝试转换为数字
                            if isinstance(value, str):
                                # 移除逗号等格式符号
                                value = value.replace(',', '').replace('，', '')
                            value = float(value)
                        except (ValueError, TypeError):
                            # 如果转换失败，保持原值
                            pass
                    # 对于字符串类型字段（如code），保持原始字符串格式
                    elif field_name == 'code' and isinstance(value, (int, float)):
                        # 如果code字段意外是数字，转换为字符串
                        value = str(value)
                        
                    self._safe_write_cell(worksheet, current_row, col_index, value)
                        
            return data_start_row + len(data_list) - 1
                        
        except Exception as e:
            self.logger.error(f"写入数据失败: {e}")
            return data_start_row
    
    def _write_total_row_safe(self, worksheet, data_list: List[Dict[str, Any]], total_row: int) -> None:
        """
        安全地写入合计行（处理合并单元格）
        
        Args:
            worksheet: 工作表对象
            data_list: 数据列表
            total_row: 合计行的行号
        """
        try:
            total_label = TARGET_CONFIG.get('total_label', '合计')
            total_label_column = TARGET_CONFIG.get('total_label_column', 1)
            total_amount_column = TARGET_CONFIG.get('total_amount_column', 4)
            
            # 确保合计行不会覆盖明细数据
            data_start_row = TARGET_CONFIG.get('data_start_row', 3)
            actual_total_row = max(total_row, data_start_row + len(data_list))
            
            # 写入合计标签
            self._safe_write_cell(worksheet, actual_total_row, total_label_column, total_label)
            
            # 计算并写入金额合计
            total_amount = 0
            for data_row in data_list:
                amount = data_row.get('amount')
                if amount is not None:
                    try:
                        # 尝试转换为数字
                        if isinstance(amount, str):
                            # 移除逗号等格式符号
                            amount = amount.replace(',', '').replace('，', '')
                        total_amount += float(amount)
                    except (ValueError, TypeError):
                        pass
            
            self._safe_write_cell(worksheet, actual_total_row, total_amount_column, total_amount)
            
            self.logger.info(f"写入合计行到第 {actual_total_row} 行: 合计金额 {total_amount}")
            
        except Exception as e:
            self.logger.error(f"写入合计行失败: {e}") 