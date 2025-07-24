# -*- coding: utf-8 -*-
"""
老师分组Excel读取器模型
负责从Excel文件中提取老师分组相关数据
"""

import os
import logging
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False
    logging.warning("xlrd库未安装，无法处理.xls文件")

from config.teacher_splitter_settings import TEACHER_SOURCE_CONFIG
from config.settings import SUPPORTED_EXTENSIONS


class TeacherExcelReader:
    """老师分组Excel文件读取器"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
    def validate_file(self, file_path: str) -> bool:
        """
        验证文件是否有效
        
        Args:
            file_path: 文件路径
            
        Returns:
            bool: 文件是否有效
        """
        try:
            if not os.path.exists(file_path):
                self.logger.error(f"文件不存在: {file_path}")
                return False
                
            _, ext = os.path.splitext(file_path)
            if ext.lower() not in SUPPORTED_EXTENSIONS:
                self.logger.error(f"不支持的文件格式: {ext}")
                return False
            
            if ext.lower() == '.xls':
                return self._validate_xls_file(file_path)
            else:
                return self._validate_xlsx_file(file_path)
                
        except Exception as e:
            self.logger.error(f"验证文件时发生错误: {e}")
            return False
    
    def _validate_xls_file(self, file_path: str) -> bool:
        """验证.xls文件"""
        try:
            if not XLRD_AVAILABLE:
                self.logger.error("xlrd库未安装，无法处理.xls文件")
                return False
                
            workbook = xlrd.open_workbook(file_path)
            return True
            
        except Exception as e:
            self.logger.error(f"无效的XLS文件 {file_path}: {e}")
            return False
    
    def _validate_xlsx_file(self, file_path: str) -> bool:
        """验证.xlsx文件"""
        try:
            workbook = load_workbook(file_path, read_only=True)
            workbook.close()
            return True
            
        except InvalidFileException as e:
            self.logger.error(f"无效的XLSX文件 {file_path}: {e}")
            return False
        except PermissionError:
            self.logger.error(f"无权限访问文件: {file_path}")
            return False

    def read_teacher_data(self, file_path: str, worksheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        从Excel文件中读取老师分组数据
        
        Args:
            file_path: Excel文件路径
            worksheet_name: 工作表名称，None表示使用第一个工作表
            
        Returns:
            List[Dict]: 提取的数据列表
        """
        try:
            if not self.validate_file(file_path):
                return []
            
            _, ext = os.path.splitext(file_path)
            if ext.lower() == '.xls':
                return self._read_xls_teacher_data(file_path, worksheet_name)
            else:
                return self._read_xlsx_teacher_data(file_path, worksheet_name)
                
        except Exception as e:
            self.logger.error(f"读取老师分组Excel文件失败: {e}")
            return []

    def _read_xlsx_teacher_data(self, file_path: str, worksheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """读取.xlsx文件的老师数据"""
        try:
            workbook = load_workbook(file_path, read_only=True)
            
            if worksheet_name and worksheet_name in workbook.sheetnames:
                worksheet = workbook[worksheet_name]
            else:
                worksheet = workbook.active
                
            self.logger.info(f"开始读取老师分组XLSX文件: {file_path}, 工作表: {worksheet.title}")
            
            data_list = []
            start_row = TEACHER_SOURCE_CONFIG['start_row']
            max_rows = TEACHER_SOURCE_CONFIG['max_rows']
            fields = TEACHER_SOURCE_CONFIG['fields']
            end_markers = TEACHER_SOURCE_CONFIG.get('end_markers', [])
            skip_empty_rows = TEACHER_SOURCE_CONFIG.get('skip_empty_rows', True)
            required_fields = TEACHER_SOURCE_CONFIG.get('required_fields', [])
            
            current_row = start_row
            
            while current_row <= min(worksheet.max_row, start_row + max_rows - 1):
                row_data = {}
                has_data = False
                is_end_marker = False
                
                # 检查是否遇到结束标识
                for field_name, col_index in fields.items():
                    cell_value = worksheet.cell(row=current_row, column=col_index).value
                    
                    if cell_value is not None:
                        cell_str = str(cell_value).strip()
                        if cell_str in end_markers:
                            is_end_marker = True
                            self.logger.info(f"遇到结束标识 '{cell_str}' 在第 {current_row} 行，停止读取")
                            break
                
                if is_end_marker:
                    break
                
                # 读取每个字段的数据
                for field_name, col_index in fields.items():
                    cell_value = worksheet.cell(row=current_row, column=col_index).value
                    
                    if cell_value is not None:
                        if isinstance(cell_value, str):
                            cell_value = cell_value.strip()
                            if cell_value:
                                has_data = True
                        else:
                            has_data = True
                    
                    row_data[field_name] = cell_value
                
                # 检查是否为空行
                if skip_empty_rows:
                    if required_fields:
                        required_has_data = any(
                            row_data.get(field) is not None and 
                            str(row_data.get(field)).strip() != ''
                            for field in required_fields
                        )
                        if not required_has_data:
                            current_row += 1
                            continue
                    elif not has_data:
                        current_row += 1
                        continue
                
                # 添加有效数据行
                if has_data or (required_fields and any(row_data.get(field) for field in required_fields)):
                    row_data['_source_file'] = os.path.basename(file_path)
                    row_data['_source_row'] = current_row
                    data_list.append(row_data)
                
                current_row += 1
                    
            workbook.close()
            self.logger.info(f"成功读取 {len(data_list)} 行老师分组数据")
            return data_list
            
        except Exception as e:
            self.logger.error(f"读取XLSX老师分组文件失败: {e}")
            return []

    def _read_xls_teacher_data(self, file_path: str, worksheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """读取.xls文件的老师数据"""
        try:
            workbook = xlrd.open_workbook(file_path)
            
            if worksheet_name and worksheet_name in workbook.sheet_names():
                worksheet = workbook.sheet_by_name(worksheet_name)
            else:
                worksheet = workbook.sheet_by_index(0)
                
            self.logger.info(f"开始读取老师分组XLS文件: {file_path}, 工作表: {worksheet.name}")
            
            data_list = []
            start_row = TEACHER_SOURCE_CONFIG['start_row'] - 1  # xlrd使用0-based索引
            max_rows = TEACHER_SOURCE_CONFIG['max_rows']
            fields = TEACHER_SOURCE_CONFIG['fields']
            end_markers = TEACHER_SOURCE_CONFIG.get('end_markers', [])
            skip_empty_rows = TEACHER_SOURCE_CONFIG.get('skip_empty_rows', True)
            required_fields = TEACHER_SOURCE_CONFIG.get('required_fields', [])
            
            current_row = start_row
            
            while current_row < min(worksheet.nrows, start_row + max_rows):
                row_data = {}
                has_data = False
                is_end_marker = False
                
                # 检查是否遇到结束标识
                for field_name, col_index in fields.items():
                    col_idx = col_index - 1  # xlrd使用0-based索引
                    if col_idx < worksheet.ncols:
                        cell_value = worksheet.cell_value(current_row, col_idx)
                        
                        if cell_value is not None:
                            cell_str = str(cell_value).strip()
                            if cell_str in end_markers:
                                is_end_marker = True
                                self.logger.info(f"遇到结束标识 '{cell_str}' 在第 {current_row + 1} 行，停止读取")
                                break
                
                if is_end_marker:
                    break
                
                # 读取每个字段的数据
                for field_name, col_index in fields.items():
                    col_idx = col_index - 1
                    cell_value = None
                    
                    if col_idx < worksheet.ncols:
                        cell_value = worksheet.cell_value(current_row, col_idx)
                        
                        if isinstance(cell_value, float) and cell_value.is_integer():
                            cell_value = int(cell_value)
                        elif isinstance(cell_value, str):
                            cell_value = cell_value.strip()
                            if cell_value:
                                has_data = True
                        elif cell_value is not None:
                            has_data = True
                    
                    row_data[field_name] = cell_value
                
                # 检查是否为空行
                if skip_empty_rows:
                    if required_fields:
                        required_has_data = any(
                            row_data.get(field) is not None and 
                            str(row_data.get(field)).strip() != ''
                            for field in required_fields
                        )
                        if not required_has_data:
                            current_row += 1
                            continue
                    elif not has_data:
                        current_row += 1
                        continue
                
                # 添加有效数据行
                if has_data or (required_fields and any(row_data.get(field) for field in required_fields)):
                    row_data['_source_file'] = os.path.basename(file_path)
                    row_data['_source_row'] = current_row + 1
                    data_list.append(row_data)
                
                current_row += 1
                    
            self.logger.info(f"成功读取 {len(data_list)} 行老师分组数据")
            return data_list
            
        except Exception as e:
            self.logger.error(f"读取XLS老师分组文件失败: {e}")
            return []

    def get_teachers_summary(self, file_path: str) -> Dict[str, Any]:
        """
        获取文件中的老师分组摘要信息
        支持多人分割（/分隔符）的正确统计
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            Dict: 老师分组摘要
        """
        try:
            data_list = self.read_teacher_data(file_path)
            
            teacher_columns = TEACHER_SOURCE_CONFIG['teacher_columns']
            summary = {
                'total_rows': len(data_list),
                'teachers': {},
                'preview': data_list[:3] if data_list else [],
            }
            
            # 统计每个角色的老师（支持多人分割）
            for role, _ in teacher_columns.items():
                teachers = set()
                for row in data_list:
                    teacher_raw = row.get(role)
                    if teacher_raw and str(teacher_raw).strip():
                        teacher_str = str(teacher_raw).strip()
                        
                        # 检查是否包含分隔符
                        if '/' in teacher_str:
                            # 多人情况：分割并添加每个老师
                            teacher_names = [name.strip() for name in teacher_str.split('/') if name.strip()]
                            for teacher_name in teacher_names:
                                teachers.add(teacher_name)
                        else:
                            # 单人情况：直接添加
                            teachers.add(teacher_str)
                            
                summary['teachers'][role] = list(teachers)
            
            return summary
            
        except Exception as e:
            self.logger.error(f"获取老师分组摘要失败: {e}")
            return {'total_rows': 0, 'teachers': {}, 'preview': []} 