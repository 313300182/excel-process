# -*- coding: utf-8 -*-
"""
工资Excel写入器
将工资数据填充到工资模板文件中
"""

import logging
import os
import json
from typing import Dict, Any, Optional, List
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from config.salary_settings import (
    SALARY_CONFIG, DEFAULT_SALARY_CONFIG, 
    JOB_SPECIFIC_CONFIG, SALARY_USER_CONFIG_FILE
)


class SalaryExcelWriter:
    """工资Excel写入器"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.template_mapping = SALARY_CONFIG['template_mapping']
        self.templates = SALARY_CONFIG['templates']
        self.user_config = self._load_user_config()
        
    def _load_user_config(self) -> Dict[str, Any]:
        """
        加载用户配置，并用默认配置补充缺失的项
        
        Returns:
            Dict[str, Any]: 用户配置
        """
        # 先获取默认配置
        config = DEFAULT_SALARY_CONFIG.copy()
        
        try:
            if os.path.exists(SALARY_USER_CONFIG_FILE):
                with open(SALARY_USER_CONFIG_FILE, 'r', encoding='utf-8') as f:
                    saved_config = json.load(f)
                    
                # 分离模板路径和配置数据
                if 'template_paths' in saved_config:
                    self.template_paths = saved_config.pop('template_paths')
                    self.logger.info(f"已加载模板路径: {list(self.template_paths.keys())}")
                
                # 深度合并配置，确保所有默认值都存在
                config = self._merge_configs(config, saved_config)
                self.logger.info("已加载用户配置并合并默认值")
                return config
        except Exception as e:
            self.logger.warning(f"加载用户配置失败: {str(e)}")
            
        return config
        
    def _merge_configs(self, default_config: Dict[str, Any], user_config: Dict[str, Any]) -> Dict[str, Any]:
        """
        深度合并配置，确保用户配置包含所有默认值
        
        Args:
            default_config: 默认配置
            user_config: 用户配置
            
        Returns:
            Dict[str, Any]: 合并后的配置
        """
        result = default_config.copy()
        
        for key, value in user_config.items():
            if key in result and isinstance(result[key], dict) and isinstance(value, dict):
                result[key] = self._merge_configs(result[key], value)
            else:
                result[key] = value
                
        return result
        
    def save_user_config(self, config: Dict[str, Any], template_paths: Dict[str, str] = None) -> bool:
        """
        保存用户配置（包括模板路径）
        
        Args:
            config: 配置数据
            template_paths: 模板文件路径
            
        Returns:
            bool: 是否保存成功
        """
        try:
            # 合并配置和模板路径
            save_config = config.copy()
            if template_paths:
                save_config['template_paths'] = template_paths
            elif hasattr(self, 'template_paths'):
                save_config['template_paths'] = self.template_paths
                
            with open(SALARY_USER_CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(save_config, f, ensure_ascii=False, indent=2)
            self.user_config = config
            self.logger.info("用户配置已保存")
            return True
        except Exception as e:
            self.logger.error(f"保存用户配置失败: {str(e)}")
            return False
            
    def set_template_paths(self, template_paths: Dict[str, str]):
        """
        设置模板文件路径
        
        Args:
            template_paths: 职业类型到模板路径的映射
        """
        self.template_paths = template_paths
        self.logger.info(f"设置模板路径: {template_paths}")
        
    def process_salary_file(self, salary_data: Dict[str, Any], 
                           job_type: str, output_dir: str) -> str:
        """
        处理工资文件，生成工资条
        
        Args:
            salary_data: 工资数据
            job_type: 职业类型
            output_dir: 输出目录
            
        Returns:
            str: 输出文件路径
        """
        workbook = None
        try:
            # 获取模板文件路径
            template_path = self._get_template_path(job_type)
            if not template_path or not os.path.exists(template_path):
                raise Exception(f"找不到 {job_type} 的模板文件: {template_path}")
            
            # 加载模板文件
            self.logger.debug(f"加载模板文件: {template_path}")
            workbook = openpyxl.load_workbook(template_path)
            worksheet = workbook.active
            
            # 验证模板格式
            self._validate_template_structure(worksheet)
            
            # 填充基本信息
            self._fill_employee_info(worksheet, salary_data['employee_info'])
            
            # 填充考勤信息
            operation_data = salary_data.get('operation_data', {})
            self._fill_attendance_info(worksheet, operation_data)
            
            # 计算和填充工资数据
            calculated_data = self._calculate_salary_data(
                salary_data, job_type)
            self._fill_salary_data(worksheet, calculated_data, job_type)
            
            # 设置所有列宽为26
            for col_num in range(1, max(worksheet.max_column, 20) + 1):
                col_letter = get_column_letter(col_num)
                worksheet.column_dimensions[col_letter].width = 26
                self.logger.debug(f"设置列 {col_letter} 宽度: 26")
            
            # 生成输出文件
            output_path = self._generate_output_path(
                salary_data['employee_info'], job_type, output_dir)
            
            # 确保输出目录存在
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # 保存文件
            workbook.save(output_path)
            self.logger.info(f"工资条已生成: {output_path}")
            
            return output_path
            
        except Exception as e:
            self.logger.error(f"处理工资文件失败: {str(e)}")
            raise Exception(f"处理工资文件失败: {str(e)}")
        finally:
            # 确保workbook被正确关闭，释放内存
            if workbook:
                try:
                    workbook.close()
                    workbook = None
                except:
                    pass
            
    def process_multiple_salary_to_single_file(self, employees_data: List[Dict[str, Any]], 
                                              output_path: str) -> str:
        """
        处理多个员工的工资数据到单个Excel文件，每个员工一个sheet
        
        Args:
            employees_data: 员工工资数据列表，每个元素包含 salary_data 和 job_type
            output_path: 输出文件路径
            
        Returns:
            str: 输出文件路径
        """
        output_workbook = None
        try:
            self.logger.info(f"开始处理 {len(employees_data)} 个员工到单个文件: {output_path}")
            
            # 创建新的工作簿
            output_workbook = openpyxl.Workbook()
            
            # 删除默认的工作表
            if 'Sheet' in [ws.title for ws in output_workbook.worksheets]:
                default_sheet = output_workbook['Sheet']
                output_workbook.remove(default_sheet)
            
            processed_count = 0
            
            for emp_data in employees_data:
                try:
                    salary_data = emp_data['salary_data']
                    job_type = emp_data['job_type']
                    employee_name = salary_data['employee_info'].get('name', '未知员工')
                    
                    self.logger.debug(f"处理员工: {employee_name}, 职业类型: {job_type}")
                    
                    # 获取模板文件路径
                    template_path = self._get_template_path(job_type)
                    if not template_path or not os.path.exists(template_path):
                        self.logger.error(f"找不到 {job_type} 的模板文件: {template_path}")
                        continue
                    
                    # 加载模板文件
                    template_workbook = openpyxl.load_workbook(template_path)
                    template_worksheet = template_workbook.active
                    
                    # 验证模板格式
                    self._validate_template_structure(template_worksheet)
                    
                    # 创建新的工作表，使用员工姓名作为sheet名
                    safe_name = self._sanitize_sheet_name(employee_name)
                    # 确保sheet名唯一
                    unique_name = self._get_unique_sheet_name(output_workbook, safe_name)
                    new_worksheet = output_workbook.create_sheet(title=unique_name)
                    
                    # 复制模板格式和内容
                    self._copy_worksheet_content(template_worksheet, new_worksheet)
                    
                    # 填充员工信息
                    self._fill_employee_info(new_worksheet, salary_data['employee_info'])
                    
                    # 填充考勤信息
                    operation_data = salary_data.get('operation_data', {})
                    self._fill_attendance_info(new_worksheet, operation_data)
                    
                    # 计算和填充工资数据
                    calculated_data = self._calculate_salary_data(salary_data, job_type)
                    self._fill_salary_data(new_worksheet, calculated_data, job_type)
                    
                    processed_count += 1
                    self.logger.debug(f"员工 {employee_name} 处理完成")
                    
                    # 关闭模板工作簿
                    template_workbook.close()
                    
                except Exception as e:
                    employee_name = emp_data.get('salary_data', {}).get('employee_info', {}).get('name', '未知员工')
                    self.logger.error(f"处理员工 {employee_name} 失败: {str(e)}")
                    continue
            
            if processed_count == 0:
                raise Exception("没有成功处理任何员工数据")
            
            # 确保输出目录存在
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # 保存文件
            output_workbook.save(output_path)
            self.logger.info(f"批量工资条已生成: {output_path}，包含 {processed_count} 个员工")
            
            return output_path
            
        except Exception as e:
            self.logger.error(f"批量处理工资文件失败: {str(e)}")
            raise Exception(f"批量处理工资文件失败: {str(e)}")
        finally:
            # 确保workbook被正确关闭，释放内存
            if output_workbook:
                try:
                    output_workbook.close()
                    output_workbook = None
                except:
                    pass
    
    def _sanitize_sheet_name(self, name: str) -> str:
        """
        清理工作表名称，移除不安全字符
        
        Args:
            name: 原名称
            
        Returns:
            str: 安全的工作表名称
        """
        # Excel工作表名称限制：不能包含 [ ] : * ? / \
        unsafe_chars = ['[', ']', ':', '*', '?', '/', '\\']
        safe_name = name
        
        for char in unsafe_chars:
            safe_name = safe_name.replace(char, '_')
        
        # 限制长度（Excel工作表名最大31个字符）
        if len(safe_name) > 31:
            safe_name = safe_name[:31]
            
        return safe_name.strip()
    
    def _get_unique_sheet_name(self, workbook, preferred_name: str) -> str:
        """
        获取唯一的工作表名称
        
        Args:
            workbook: 工作簿
            preferred_name: 首选名称
            
        Returns:
            str: 唯一的工作表名称
        """
        existing_names = [ws.title for ws in workbook.worksheets]
        
        if preferred_name not in existing_names:
            return preferred_name
        
        # 如果名称已存在，添加数字后缀
        counter = 1
        while True:
            candidate_name = f"{preferred_name}_{counter}"
            if len(candidate_name) > 31:
                # 如果加数字后超过31个字符，截短原名称
                base_length = 31 - len(f"_{counter}")
                candidate_name = f"{preferred_name[:base_length]}_{counter}"
            
            if candidate_name not in existing_names:
                return candidate_name
            counter += 1
    
    def _copy_worksheet_content(self, source_ws, target_ws):
        """
        复制工作表内容（包括格式、公式、样式等）
        
        Args:
            source_ws: 源工作表
            target_ws: 目标工作表
        """
        try:
            # 复制单元格数据和格式
            for row in source_ws.iter_rows():
                for cell in row:
                    target_cell = target_ws[cell.coordinate]
                    
                    # 复制值
                    if cell.value is not None:
                        target_cell.value = cell.value
                    
                    # 复制格式
                    if cell.has_style:
                        target_cell.font = cell.font.copy()
                        target_cell.border = cell.border.copy()
                        target_cell.fill = cell.fill.copy()
                        target_cell.number_format = cell.number_format
                        target_cell.protection = cell.protection.copy()
                        target_cell.alignment = cell.alignment.copy()
            
            # 复制合并单元格
            for merged_range in source_ws.merged_cells.ranges:
                target_ws.merge_cells(str(merged_range))
            
            # 复制行高
            for row_num in range(1, source_ws.max_row + 1):
                if row_num in source_ws.row_dimensions:
                    source_height = source_ws.row_dimensions[row_num].height
                    if source_height is not None:
                        target_ws.row_dimensions[row_num].height = source_height
                        
            # 设置所有列宽为26
            for col_num in range(1, max(source_ws.max_column, 20) + 1):
                col_letter = get_column_letter(col_num)
                target_ws.column_dimensions[col_letter].width = 26
                self.logger.debug(f"设置列 {col_letter} 宽度: 26")
                        
            self.logger.debug("工作表内容复制完成")
            
        except Exception as e:
            self.logger.warning(f"复制工作表内容时出现警告: {str(e)}")
            # 即使复制格式失败，也不影响数据填充
            
    def _get_template_path(self, job_type: str) -> Optional[str]:
        """
        获取模板文件路径
        
        Args:
            job_type: 职业类型
            
        Returns:
            Optional[str]: 模板文件路径
        """
        if hasattr(self, 'template_paths') and job_type in self.template_paths:
            return self.template_paths[job_type]
        return None
        
    def _fill_employee_info(self, worksheet, employee_info: Dict[str, Any]):
        """
        填充员工基本信息
        
        Args:
            worksheet: 工作表
            employee_info: 员工信息
        """
        try:
            # 填充姓名
            if 'name' in employee_info and employee_info['name']:
                name_cell = self.template_mapping['employee_name']
                self.logger.debug(f"写入姓名到单元格 {name_cell}: {employee_info['name']}")
                worksheet[name_cell] = str(employee_info['name'])
                
            # 填充月份
            if 'month' in employee_info and employee_info['month']:
                month_cell = self.template_mapping['month']
                self.logger.debug(f"写入月份到单元格 {month_cell}: {employee_info['month']}")
                worksheet[month_cell] = str(employee_info['month'])
                
            self.logger.debug("员工基本信息填充完成")
            
        except Exception as e:
            self.logger.error(f"填充员工信息失败: {str(e)}")
            raise Exception(f"填充员工信息失败: {str(e)}")
            
    def _fill_attendance_info(self, worksheet, operation_data: Dict[str, Any]):
        """
        填充考勤信息到模板
        
        Args:
            worksheet: 工作表
            operation_data: 操作数据（包含考勤信息）
        """
        try:
            self.logger.debug("开始填充考勤信息")
            
            # 获取考勤信息映射
            attendance_mapping = self.template_mapping.get('attendance_info', {})
            
            if not attendance_mapping:
                self.logger.warning("未找到考勤信息映射配置，跳过考勤信息填充")
                return
            
            # 填充上班天数
            if 'work_days' in attendance_mapping:
                work_days = operation_data.get('work_days', 0)
                cell = attendance_mapping['work_days']
                worksheet[cell] = int(work_days) if work_days else 0
                self.logger.debug(f"写入上班天数到 {cell}: {work_days}")
            
            # 填充休息天数
            if 'rest_days' in attendance_mapping:
                rest_days = operation_data.get('rest_days', 0)
                cell = attendance_mapping['rest_days']
                worksheet[cell] = int(rest_days) if rest_days else 0
                self.logger.debug(f"写入休息天数到 {cell}: {rest_days}")
            
            # 填充迟到次数
            if 'late_count' in attendance_mapping:
                late_count = operation_data.get('late_count', 0)
                cell = attendance_mapping['late_count']
                worksheet[cell] = int(late_count) if late_count else 0
                self.logger.debug(f"写入迟到次数到 {cell}: {late_count}")
            
            # 填充培训天数
            if 'training_days' in attendance_mapping:
                training_days = operation_data.get('training_days', 0)
                cell = attendance_mapping['training_days']
                worksheet[cell] = int(training_days) if training_days else 0
                self.logger.debug(f"写入培训天数到 {cell}: {training_days}")
            
            self.logger.debug("考勤信息填充完成")
            
        except Exception as e:
            self.logger.error(f"填充考勤信息失败: {str(e)}")
            # 考勤信息填充失败不阻止工资数据处理，只记录警告
            self.logger.warning("考勤信息填充失败，将继续处理工资数据")
            
    def _calculate_salary_data(self, salary_data: Dict[str, Any], 
                              job_type: str) -> Dict[str, Any]:
        """
        计算工资数据（基于业绩数据和手工费数据）
        
        Args:
            salary_data: 包含员工信息、业绩数据和操作数据的字典
            job_type: 职业类型
            
        Returns:
            Dict[str, Any]: 计算后的工资数据（包含数量、单价、金额）
        """
        calculated = {}
        
        try:
            employee_name = salary_data['employee_info'].get('name', '')
            performance_data = salary_data.get('performance_data', {})
            operation_data = salary_data.get('operation_data', {})
            
            # 获取配置
            base_config = self.user_config.get('base_salary', {})
            floating_config = self.user_config.get('floating_salary', {})
            commission_config = self.user_config.get('commission_rates', {})
            manual_config = self.user_config.get('manual_fees', {})

            other_config = self.user_config.get('other_config', {})
            job_config = JOB_SPECIFIC_CONFIG.get(job_type, {})
            
            # 基本底薪（数量默认为1，根据职业类型设置不同的默认值）
            # 先检查是否有该员工的特殊底薪设置
            base_salary_rate = base_config.get('special_rates', {}).get(employee_name)
            if base_salary_rate is None:
                # 如果没有特殊设置，优先从用户配置的职业特定配置中获取
                job_specific_config = self.user_config.get('job_specific_config', {})
                job_specific_data = job_specific_config.get(job_type, {})
                if job_specific_data and 'base_salary' in job_specific_data:
                    base_salary_rate = job_specific_data['base_salary']
                else:
                    # 如果用户配置中没有，使用默认的职业配置
                    job_default_salary = job_config.get('default_base_salary', 5000)
                    base_salary_rate = job_default_salary
            
            base_salary_quantity = 1  # 基本底薪数量固定为1
            calculated['base_salary_quantity'] = base_salary_quantity
            calculated['base_salary_rate'] = base_salary_rate
            calculated['base_salary'] = base_salary_quantity * base_salary_rate  # 用于汇总计算
            
            # 浮动底薪（数量默认为1）
            # 先检查是否有该员工的特殊浮动底薪设置
            floating_salary_rate = floating_config.get('special_rates', {}).get(employee_name)
            if floating_salary_rate is None:
                # 如果没有特殊设置，优先从用户配置的职业特定配置中获取
                if job_specific_data and 'floating_salary' in job_specific_data:
                    floating_salary_rate = job_specific_data['floating_salary']
                else:
                    # 如果用户配置中没有，使用通用默认值
                    floating_salary_rate = floating_config.get('default', 0)
                    
            floating_salary_quantity = 1  # 浮动底薪数量固定为1
            calculated['floating_salary_quantity'] = floating_salary_quantity
            calculated['floating_salary_rate'] = floating_salary_rate
            calculated['floating_salary'] = floating_salary_quantity * floating_salary_rate
            
            # 计算提成（基于业绩数据和职业类型）
            performance_value = performance_data.get('total_performance_value', 0)
            commission_new_config = self.user_config.get('commission_config', {})
            
            # 初始化所有提成项目为0
            calculated['expert_commission_quantity'] = 0
            calculated['expert_commission_rate'] = 0
            calculated['expert_commission'] = 0
            calculated['service_commission_quantity'] = 0
            calculated['service_commission_rate'] = 0
            calculated['service_commission'] = 0
            calculated['operation_commission_quantity'] = 0
            calculated['operation_commission_rate'] = 0
            calculated['operation_commission'] = 0
            
            # 根据职业类型计算对应的提成
            if job_type == '服务总监':
                # 专家提成
                expert_config = commission_new_config.get('expert_commission', {})
                calculated['expert_commission_quantity'] = performance_value
                calculated['expert_commission_rate'] = expert_config.get('default_rate', 1.2)
                # 提成金额将通过Excel公式计算：=数量×(比例÷100)
                calculated['expert_commission'] = performance_value * (calculated['expert_commission_rate'] / 100)
                
            elif job_type == '服务老师':
                # 服务提成
                service_config = commission_new_config.get('service_commission', {})
                calculated['service_commission_quantity'] = performance_value
                calculated['service_commission_rate'] = service_config.get('default_rate', 
                    commission_config.get('service_rate', 1.5))
                # 提成金额 = 数量 × (比例 ÷ 100)，配置中的1.5表示1.5%
                calculated['service_commission'] = performance_value * (calculated['service_commission_rate'] / 100)
                
            elif job_type == '操作老师':
                # 操作提成
                operation_config = commission_new_config.get('operation_commission', {})
                calculated['operation_commission_quantity'] = performance_value
                calculated['operation_commission_rate'] = operation_config.get('default_rate',
                    commission_config.get('operation_rate', 0.8))
                # 提成金额 = 数量 × (比例 ÷ 100)，配置中的0.8表示0.8%
                calculated['operation_commission'] = performance_value * (calculated['operation_commission_rate'] / 100)
            
            # 培训补贴（数量从操作表Sheet2获取）
            training_allowance_rate = other_config.get('training_allowance', 0)
            training_allowance_quantity = operation_data.get('training_days', 0)  # 从操作表获取培训天数
            calculated['training_allowance_quantity'] = training_allowance_quantity
            calculated['training_allowance_rate'] = training_allowance_rate
            calculated['training_allowance'] = training_allowance_quantity * training_allowance_rate
            
            # 身体部位手工费（从操作表获取数量）
            body_manual_fee_quantity = operation_data.get('body_count', 0)
            body_manual_fee_rate = manual_config.get('body_rate', 0)
            calculated['body_manual_fee_quantity'] = body_manual_fee_quantity
            calculated['body_manual_fee_rate'] = body_manual_fee_rate
            calculated['body_manual_fee'] = body_manual_fee_quantity * body_manual_fee_rate
            
            # 面部手工费（从操作表获取数量）
            face_manual_fee_quantity = operation_data.get('face_count', 0)
            face_manual_fee_rate = manual_config.get('face_rate', 0)
            calculated['face_manual_fee_quantity'] = face_manual_fee_quantity
            calculated['face_manual_fee_rate'] = face_manual_fee_rate
            calculated['face_manual_fee'] = face_manual_fee_quantity * face_manual_fee_rate
            
            # 特殊补贴（职业特有）
            calculated['special_allowance'] = job_config.get('special_allowance', 0)
            
            # 计算应发合计
            salary_total = (
                calculated['base_salary'] +
                calculated['floating_salary'] +
                calculated['expert_commission'] +
                calculated['service_commission'] +
                calculated['operation_commission'] +
                calculated['training_allowance'] +
                calculated['body_manual_fee'] +
                calculated['face_manual_fee'] +
                calculated['special_allowance']
            )
            calculated['total_salary'] = salary_total
            
            # 计算缺勤扣减（数量和单价）
            absent_days = operation_data.get('actual_absent_days', 0)  # 实际缺勤天数
            base_salary_amount = calculated['base_salary']  # 基本底薪金额
            
            # 获取当前月份天数
            from datetime import datetime
            import calendar
            now = datetime.now()
            current_month_days = calendar.monthrange(now.year, now.month)[1]
            
            # 缺勤单价 = 当月底薪/当月天数（正数）
            absent_deduction_rate = (base_salary_amount / current_month_days) if current_month_days > 0 else 0
            # 缺勤数量 = -实际缺勤天数（只有实际缺勤>0时才扣减）
            absent_deduction_quantity = -absent_days if absent_days > 0 else 0
            
            calculated['absent_deduction_quantity'] = absent_deduction_quantity
            calculated['absent_deduction_rate'] = absent_deduction_rate
            calculated['absent_deduction'] = absent_deduction_quantity * absent_deduction_rate
            
            # 计算迟到扣减（从配置获取迟到单价）
            late_count = operation_data.get('late_count', 0)
            late_deduction_rate = self.user_config.get('other_config', {}).get('late_deduction_rate', 20)  # 默认20（正数）
            # 迟到数量 = -实际迟到次数（负数表示扣减）
            late_deduction_quantity = -late_count if late_count > 0 else 0
            
            calculated['late_deduction_quantity'] = late_deduction_quantity
            calculated['late_deduction_rate'] = late_deduction_rate
            calculated['late_deduction'] = late_deduction_quantity * late_deduction_rate
            
            # 社保（数量默认-1，单价从配置获取）
            social_security_rate = other_config.get('social_security_rate', 505.26)
            calculated['social_security_quantity'] = -1
            calculated['social_security_rate'] = social_security_rate
            calculated['social_security'] = calculated['social_security_quantity'] * calculated['social_security_rate']
            
            self.logger.info(f"💰 社保扣除: 数量={calculated['social_security_quantity']}, 单价={social_security_rate:.2f}元, 金额={calculated['social_security']:.2f}元")
            
            # 个人所得税（数量默认-1，单价从操作表Sheet2的F列获取）
            personal_tax_rate = operation_data.get('personal_tax_amount', 0)
            self.logger.info(f"🔍 从操作表获取个人所得税: personal_tax_amount={personal_tax_rate}, operation_data keys: {list(operation_data.keys())}")
            
            calculated['personal_tax_quantity'] = -1
            calculated['personal_tax_rate'] = personal_tax_rate
            calculated['personal_tax'] = calculated['personal_tax_quantity'] * calculated['personal_tax_rate']
            
            self.logger.info(f"💰 个人所得税: 数量={calculated['personal_tax_quantity']}, 单价={personal_tax_rate:.2f}元, 金额={calculated['personal_tax']:.2f}元")
            
            # 扣减小计
            deduction_total = (
                calculated['absent_deduction'] +
                calculated['late_deduction'] +
                calculated['social_security'] +
                calculated['personal_tax']
            )
            calculated['total_deduction'] = deduction_total
            
            # 实发工资（扣减项目为负数，所以用加法）
            calculated['net_salary'] = salary_total + deduction_total
            
            self.logger.info(f"员工 {employee_name} 工资计算完成，应发: {salary_total:.2f}, 实发: {calculated['net_salary']:.2f}")
            
        except Exception as e:
            self.logger.error(f"计算工资数据失败: {str(e)}")
            
        return calculated
        
    def _extract_quantity_from_details(self, details: List[Dict[str, Any]], 
                                     keyword: str) -> float:
        """
        从明细中提取数量
        
        Args:
            details: 工资明细
            keyword: 关键词
            
        Returns:
            float: 数量
        """
        total_quantity = 0.0
        
        try:
            for detail in details:
                project = str(detail.get('project', ''))
                if keyword in project:
                    quantity = detail.get('quantity', 0) or 0
                    total_quantity += float(quantity)
                    
        except Exception as e:
            self.logger.warning(f"提取 {keyword} 数量时出错: {str(e)}")
            
        return total_quantity
        
    def _fill_salary_data(self, worksheet, calculated_data: Dict[str, Any], job_type: str = None):
        """
        填充工资数据到模板（使用数量、单价和公式）
        
        Args:
            worksheet: 工作表
            calculated_data: 计算后的工资数据
        """
        try:
            self.logger.debug(f"开始填充工资数据（新格式：数量×单价=金额），共 {len(calculated_data)} 项")
            
            # 获取应发项目映射
            salary_mapping = self.template_mapping.get('salary_items', {})
            self.logger.debug(f"应发项目映射: {salary_mapping}")
            
            # 根据职业类型确定需要处理的工资项目（不包括应发合计）
            salary_items = [
                'base_salary',          # 基本底薪
                'floating_salary',      # 浮动底薪
                'training_allowance',   # 培训补贴
                'body_manual_fee',      # 身体部位手工费
                'face_manual_fee'       # 面部手工费
            ]
            
            # 根据职业类型添加对应的提成项目
            if job_type == '服务总监':
                salary_items.append('expert_commission')      # 专家提成
            elif job_type == '服务老师':
                salary_items.append('service_commission')     # 服务提成
            elif job_type == '操作老师':
                salary_items.append('operation_commission')   # 操作提成
            
            # 填充每个工资项目的数量、单价和公式
            for item in salary_items:
                quantity_key = f"{item}_quantity"
                rate_key = f"{item}_rate"
                amount_key = item
                
                # 填充数量（包括0值）
                if quantity_key in salary_mapping and quantity_key in calculated_data:
                    quantity_cell = salary_mapping[quantity_key]
                    quantity_value = calculated_data[quantity_key]
                    # 数量填入整数
                    worksheet[quantity_cell] = int(quantity_value) if quantity_value is not None else 0
                    self.logger.debug(f"写入{item}数量到 {quantity_cell}: {quantity_value}")
                
                # 填充单价（包括0值）
                if rate_key in salary_mapping and rate_key in calculated_data:
                    rate_cell = salary_mapping[rate_key]
                    rate_value = calculated_data[rate_key]
                    
                    # 对于提成比例，需要转换为小数并设置百分比格式
                    if item in ['expert_commission', 'service_commission', 'operation_commission']:
                        # 将1.5转换为0.015，显示为1.5%
                        decimal_value = round(float(rate_value) / 100, 4) if rate_value is not None else 0
                        worksheet[rate_cell] = decimal_value
                        worksheet[rate_cell].number_format = '0.00%'
                        self.logger.debug(f"写入{item}比例到 {rate_cell}: {rate_value}% (小数值: {decimal_value})")
                    else:
                        # 其他项目单价填入保留2位小数的数值
                        worksheet[rate_cell] = round(float(rate_value), 2) if rate_value is not None else 0
                        self.logger.debug(f"写入{item}单价到 {rate_cell}: {rate_value}")
                
                # 填充金额公式
                if (amount_key in salary_mapping and 
                    quantity_key in salary_mapping and 
                    rate_key in salary_mapping and
                    quantity_key in calculated_data and 
                    rate_key in calculated_data):
                    
                    amount_cell = salary_mapping[amount_key]
                    quantity_cell = salary_mapping[quantity_key]
                    rate_cell = salary_mapping[rate_key]
                    
                    # 所有项目统一使用：=数量*单价
                    formula = f"={quantity_cell}*{rate_cell}"
                    
                    worksheet[amount_cell] = formula
                    self.logger.debug(f"写入{item}金额公式到 {amount_cell}: {formula}")
            
            # 填充应发合计公式（所有应发项目金额之和）
            if 'total_salary' in salary_mapping:
                total_cell = salary_mapping['total_salary']
                # 构建求和公式
                amount_cells = []
                for item in salary_items:
                    if item in salary_mapping:
                        amount_cells.append(salary_mapping[item])
                
                if amount_cells:
                    formula = f"=SUM({','.join(amount_cells)})"
                    worksheet[total_cell] = formula
                    self.logger.debug(f"写入应发合计公式到 {total_cell}: {formula}")
                        
            # 填充扣减项目（包含数量、单价和公式）
            deduction_mapping = self.template_mapping.get('deduction_items', {})
            self.logger.debug(f"填充扣减项目，映射: {deduction_mapping}")
            
            # 定义需要处理的扣减项目（带数量和单价的）
            deduction_items_with_quantity = ['absent_deduction', 'late_deduction', 'social_security', 'personal_tax']
            
            # 填充扣减项目的数量、单价和公式
            for item in deduction_items_with_quantity:
                quantity_key = f"{item}_quantity"
                rate_key = f"{item}_rate"
                amount_key = item
                
                # 填充数量（包括0值）
                if quantity_key in deduction_mapping and quantity_key in calculated_data:
                    quantity_cell = deduction_mapping[quantity_key]
                    quantity_value = calculated_data[quantity_key]
                    worksheet[quantity_cell] = int(quantity_value) if quantity_value is not None else 0
                    self.logger.debug(f"写入{item}数量到 {quantity_cell}: {quantity_value}")
                
                # 填充单价（包括0值）
                if rate_key in deduction_mapping and rate_key in calculated_data:
                    rate_cell = deduction_mapping[rate_key]
                    rate_value = calculated_data[rate_key]
                    worksheet[rate_cell] = round(float(rate_value), 2) if rate_value is not None else 0
                    self.logger.debug(f"写入{item}单价到 {rate_cell}: {rate_value}")
                
                # 填充金额公式
                if (amount_key in deduction_mapping and 
                    quantity_key in deduction_mapping and 
                    rate_key in deduction_mapping and
                    quantity_key in calculated_data and 
                    rate_key in calculated_data):
                    
                    amount_cell = deduction_mapping[amount_key]
                    quantity_cell = deduction_mapping[quantity_key]
                    rate_cell = deduction_mapping[rate_key]
                    
                    # 创建Excel公式：=数量格*单价格
                    formula = f"={quantity_cell}*{rate_cell}"
                    worksheet[amount_cell] = formula
                    self.logger.debug(f"写入{item}金额公式到 {amount_cell}: {formula}")
            

            # 填充扣减小计公式
            if 'total_deduction' in deduction_mapping:
                deduction_cells = []
                for deduction_key in ['absent_deduction', 'late_deduction', 'social_security', 'personal_tax']:
                    if deduction_key in deduction_mapping:
                        deduction_cells.append(deduction_mapping[deduction_key])
                
                if deduction_cells:
                    total_cell = deduction_mapping['total_deduction']
                    formula = f"=SUM({','.join(deduction_cells)})"
                    worksheet[total_cell] = formula
                    self.logger.debug(f"写入扣减小计公式到 {total_cell}: {formula}")
                        
            # 填充实发工资公式
            net_cell = self.template_mapping.get('net_salary')
            total_salary_cell = salary_mapping.get('total_salary')
            total_deduction_cell = deduction_mapping.get('total_deduction')
            
            if net_cell and total_salary_cell and total_deduction_cell:
                try:
                    # 实发工资 = 应发合计 + 扣减小计（扣减项目为负数）
                    formula = f"={total_salary_cell}+{total_deduction_cell}"
                    worksheet[net_cell] = formula
                    self.logger.debug(f"写入实发工资公式到 {net_cell}: {formula}")
                except Exception as e:
                    self.logger.error(f"写入实发工资公式到 {net_cell} 失败: {str(e)}")
                    raise
            
            self.logger.debug("工资数据填充完成（使用数量×单价公式）")
            
        except Exception as e:
            self.logger.error(f"填充工资数据失败: {str(e)}")
            raise Exception(f"填充工资数据失败: {str(e)}")
            
    def _generate_output_path(self, employee_info: Dict[str, Any], 
                            job_type: str, output_dir: str) -> str:
        """
        生成输出文件路径
        
        Args:
            employee_info: 员工信息
            job_type: 职业类型
            output_dir: 输出目录
            
        Returns:
            str: 输出文件路径
        """
        name = employee_info.get('name', '未知员工')
        month = employee_info.get('month', '未知月份')
        
        # 清理文件名中的特殊字符
        safe_name = self._sanitize_filename(name)
        safe_month = self._sanitize_filename(month)
        safe_job_type = self._sanitize_filename(job_type)
        
        filename = f"{safe_name}_{safe_month}_{safe_job_type}_工资条.xlsx"
        return os.path.join(output_dir, filename)
        
    def get_user_config(self) -> Dict[str, Any]:
        """
        获取当前用户配置
        
        Returns:
            Dict[str, Any]: 用户配置
        """
        return self.user_config.copy()
        
    def validate_templates(self, template_paths: Dict[str, str]) -> Dict[str, bool]:
        """
        验证模板文件
        
        Args:
            template_paths: 模板文件路径
            
        Returns:
            Dict[str, bool]: 验证结果
        """
        results = {}
        
        for job_type, template_path in template_paths.items():
            try:
                if not os.path.exists(template_path):
                    results[job_type] = False
                    continue
                    
                # 尝试打开文件
                workbook = openpyxl.load_workbook(template_path)
                worksheet = workbook.active
                
                # 检查关键单元格是否存在
                required_cells = [
                    self.template_mapping['employee_name'],
                    self.template_mapping['month'],
                    self.template_mapping['net_salary']
                ]
                
                valid = True
                for cell in required_cells:
                    try:
                        _ = worksheet[cell]
                    except:
                        valid = False
                        break
                        
                results[job_type] = valid
                
            except Exception as e:
                self.logger.error(f"验证模板 {job_type} 失败: {str(e)}")
                results[job_type] = False
                
        return results 
        
    def _validate_template_structure(self, worksheet):
        """
        验证模板结构是否正确
        
        Args:
            worksheet: 工作表
        """
        try:
            self.logger.debug("开始验证模板结构")
            
            # 验证关键单元格是否存在
            critical_cells = [
                self.template_mapping.get('employee_name'),
                self.template_mapping.get('month'),
                self.template_mapping.get('net_salary')
            ]
            
            for cell in critical_cells:
                if cell:
                    try:
                        # 尝试访问单元格
                        _ = worksheet[cell]
                        self.logger.debug(f"验证单元格 {cell}: OK")
                    except Exception as e:
                        raise Exception(f"无法访问关键单元格 {cell}: {str(e)}")
            
            # 验证工资项目单元格
            salary_items = self.template_mapping.get('salary_items', {})
            for key, cell in salary_items.items():
                try:
                    _ = worksheet[cell]
                    self.logger.debug(f"验证工资项目单元格 {key}({cell}): OK")
                except Exception as e:
                    self.logger.warning(f"工资项目单元格 {key}({cell}) 访问异常: {str(e)}")
            
            self.logger.debug("模板结构验证完成")
            
        except Exception as e:
            self.logger.error(f"模板结构验证失败: {str(e)}")
            raise Exception(f"模板结构验证失败: {str(e)}") 