# -*- coding: utf-8 -*-
"""
Excel读取器模型
负责从Excel文件中按配置提取数据
"""

import os
import logging
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False
    logging.warning("xlrd库未安装，无法处理.xls文件")

from config.settings import SOURCE_CONFIG, SUPPORTED_EXTENSIONS


class ExcelReader:
    """Excel文件读取器"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
    def validate_file(self, file_path: str) -> bool:
        """
        验证文件是否有效
        
        Args:
            file_path: 文件路径
            
        Returns:
            bool: 文件是否有效
        """
        try:
            # 检查文件是否存在
            if not os.path.exists(file_path):
                self.logger.error(f"文件不存在: {file_path}")
                return False
                
            # 检查文件扩展名
            _, ext = os.path.splitext(file_path)
            if ext.lower() not in SUPPORTED_EXTENSIONS:
                self.logger.error(f"不支持的文件格式: {ext}")
                return False
            
            # 根据文件扩展名选择不同的验证方法
            if ext.lower() == '.xls':
                return self._validate_xls_file(file_path)
            else:
                return self._validate_xlsx_file(file_path)
                
        except Exception as e:
            self.logger.error(f"验证文件时发生错误: {e}")
            return False
    
    def _validate_xls_file(self, file_path: str) -> bool:
        """验证.xls文件"""
        try:
            if not XLRD_AVAILABLE:
                self.logger.error("xlrd库未安装，无法处理.xls文件。请运行: pip install xlrd==2.0.1")
                return False
                
            workbook = xlrd.open_workbook(file_path)
            return True
            
        except Exception as e:
            self.logger.error(f"无效的XLS文件 {file_path}: {e}")
            return False
    
    def _validate_xlsx_file(self, file_path: str) -> bool:
        """验证.xlsx文件"""
        try:
            workbook = load_workbook(file_path, read_only=True)
            workbook.close()
            return True
            
        except InvalidFileException as e:
            self.logger.error(f"无效的XLSX文件 {file_path}: {e}")
            return False
        except PermissionError:
            self.logger.error(f"无权限访问文件: {file_path}")
            return False
    
    def read_data(self, file_path: str, worksheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        从Excel文件中读取数据（支持多行明细数据）
        
        Args:
            file_path: Excel文件路径
            worksheet_name: 工作表名称，None表示使用第一个工作表
            
        Returns:
            List[Dict]: 提取的数据列表
        """
        try:
            if not self.validate_file(file_path):
                return []
            
            # 根据文件扩展名选择不同的读取方法
            _, ext = os.path.splitext(file_path)
            if ext.lower() == '.xls':
                return self._read_xls_data(file_path, worksheet_name)
            else:
                return self._read_xlsx_data(file_path, worksheet_name)
                
        except Exception as e:
            self.logger.error(f"读取Excel文件失败: {e}")
            return []
    
    def _read_xls_data(self, file_path: str, worksheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """读取.xls文件数据"""
        try:
            workbook = xlrd.open_workbook(file_path)
            
            # 选择工作表
            if worksheet_name and worksheet_name in workbook.sheet_names():
                worksheet = workbook.sheet_by_name(worksheet_name)
            else:
                worksheet = workbook.sheet_by_index(0)
                
            self.logger.info(f"开始读取XLS文件: {file_path}, 工作表: {worksheet.name}")
            
            data_list = []
            start_row = SOURCE_CONFIG['start_row'] - 1  # xlrd使用0-based索引
            max_rows = SOURCE_CONFIG['max_rows']
            fields = SOURCE_CONFIG['fields']
            end_markers = SOURCE_CONFIG.get('end_markers', [])
            skip_empty_rows = SOURCE_CONFIG.get('skip_empty_rows', True)
            required_fields = SOURCE_CONFIG.get('required_fields', [])
            
            # 逐行读取数据
            current_row = start_row
            
            while current_row < min(worksheet.nrows, start_row + max_rows):
                row_data = {}
                has_data = False
                is_end_marker = False
                
                # 检查是否遇到结束标识
                for field_name, col_index in fields.items():
                    col_idx = col_index - 1  # xlrd使用0-based索引
                    if col_idx < worksheet.ncols:
                        cell_value = worksheet.cell_value(current_row, col_idx)
                        
                        # 转换为字符串进行比较
                        if cell_value is not None:
                            cell_str = str(cell_value).strip()
                            if cell_str in end_markers:
                                is_end_marker = True
                                self.logger.info(f"遇到结束标识 '{cell_str}' 在第 {current_row + 1} 行，停止读取")
                                break
                
                # 如果遇到结束标识，停止读取
                if is_end_marker:
                    break
                
                # 读取每个字段的数据
                for field_name, col_index in fields.items():
                    col_idx = col_index - 1  # xlrd使用0-based索引
                    cell_value = None
                    
                    if col_idx < worksheet.ncols:
                        cell_value = worksheet.cell_value(current_row, col_idx)
                        
                        # 处理xlrd的数据类型
                        if isinstance(cell_value, float) and cell_value.is_integer():
                            cell_value = int(cell_value)
                        elif isinstance(cell_value, str):
                            cell_value = cell_value.strip()
                            if cell_value:
                                has_data = True
                        elif cell_value is not None:
                            has_data = True
                    
                    row_data[field_name] = cell_value
                
                # 检查是否为空行
                if skip_empty_rows:
                    if required_fields:
                        required_has_data = any(
                            row_data.get(field) is not None and 
                            str(row_data.get(field)).strip() != ''
                            for field in required_fields
                        )
                        if not required_has_data:
                            self.logger.debug(f"跳过空行: 第 {current_row + 1} 行")
                            current_row += 1
                            continue
                    elif not has_data:
                        self.logger.debug(f"跳过空行: 第 {current_row + 1} 行")
                        current_row += 1
                        continue
                
                # 添加有效数据行
                if has_data or (required_fields and any(row_data.get(field) for field in required_fields)):
                    row_data['_source_file'] = os.path.basename(file_path)
                    row_data['_source_row'] = current_row + 1  # 转换回1-based
                    data_list.append(row_data)
                    self.logger.debug(f"读取第 {current_row + 1} 行数据: {row_data}")
                
                current_row += 1
                    
            self.logger.info(f"成功读取 {len(data_list)} 行数据")
            return data_list
            
        except Exception as e:
            self.logger.error(f"读取XLS文件失败: {e}")
            return []
    
    def _read_xlsx_data(self, file_path: str, worksheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """读取.xlsx文件数据"""
        try:
            workbook = load_workbook(file_path, read_only=True)
            
            # 选择工作表
            if worksheet_name and worksheet_name in workbook.sheetnames:
                worksheet = workbook[worksheet_name]
            else:
                worksheet = workbook.active
                
            self.logger.info(f"开始读取XLSX文件: {file_path}, 工作表: {worksheet.title}")
            
            data_list = []
            start_row = SOURCE_CONFIG['start_row']
            max_rows = SOURCE_CONFIG['max_rows']
            fields = SOURCE_CONFIG['fields']
            end_markers = SOURCE_CONFIG.get('end_markers', [])
            skip_empty_rows = SOURCE_CONFIG.get('skip_empty_rows', True)
            required_fields = SOURCE_CONFIG.get('required_fields', [])
            
            # 逐行读取数据
            current_row = start_row
            
            while current_row <= min(worksheet.max_row, start_row + max_rows - 1):
                row_data = {}
                has_data = False
                is_end_marker = False
                
                # 检查是否遇到结束标识
                for field_name, col_index in fields.items():
                    cell_value = worksheet.cell(row=current_row, column=col_index).value
                    
                    # 转换为字符串进行比较
                    if cell_value is not None:
                        cell_str = str(cell_value).strip()
                        if cell_str in end_markers:
                            is_end_marker = True
                            self.logger.info(f"遇到结束标识 '{cell_str}' 在第 {current_row} 行，停止读取")
                            break
                
                # 如果遇到结束标识，停止读取
                if is_end_marker:
                    break
                
                # 读取每个字段的数据
                for field_name, col_index in fields.items():
                    cell_value = worksheet.cell(row=current_row, column=col_index).value
                    
                    # 清理数据
                    if cell_value is not None:
                        # 处理字符串类型
                        if isinstance(cell_value, str):
                            cell_value = cell_value.strip()
                            if cell_value:  # 非空字符串
                                has_data = True
                        else:
                            has_data = True
                    
                    row_data[field_name] = cell_value
                
                # 检查是否为空行
                if skip_empty_rows:
                    # 检查必填字段是否都为空
                    if required_fields:
                        required_has_data = any(
                            row_data.get(field) is not None and 
                            str(row_data.get(field)).strip() != ''
                            for field in required_fields
                        )
                        if not required_has_data:
                            self.logger.debug(f"跳过空行: 第 {current_row} 行")
                            current_row += 1
                            continue
                    elif not has_data:
                        self.logger.debug(f"跳过空行: 第 {current_row} 行")
                        current_row += 1
                        continue
                
                # 添加有效数据行
                if has_data or (required_fields and any(row_data.get(field) for field in required_fields)):
                    row_data['_source_file'] = os.path.basename(file_path)
                    row_data['_source_row'] = current_row
                    data_list.append(row_data)
                    self.logger.debug(f"读取第 {current_row} 行数据: {row_data}")
                
                current_row += 1
                    
            workbook.close()
            self.logger.info(f"成功读取 {len(data_list)} 行数据")
            return data_list
            
        except Exception as e:
            self.logger.error(f"读取XLSX文件失败: {e}")
            return []
    
    def get_worksheet_names(self, file_path: str) -> List[str]:
        """
        获取Excel文件中的所有工作表名称
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            List[str]: 工作表名称列表
        """
        try:
            if not self.validate_file(file_path):
                return []
            
            _, ext = os.path.splitext(file_path)
            if ext.lower() == '.xls':
                workbook = xlrd.open_workbook(file_path)
                return workbook.sheet_names()
            else:
                workbook = load_workbook(file_path, read_only=True)
                sheet_names = workbook.sheetnames
                workbook.close()
                return sheet_names
            
        except Exception as e:
            self.logger.error(f"获取工作表名称失败: {e}")
            return []
    
    def preview_data(self, file_path: str, rows: int = 10) -> List[Dict[str, Any]]:
        """
        预览Excel文件的前几行数据
        
        Args:
            file_path: Excel文件路径
            rows: 预览行数
            
        Returns:
            List[Dict]: 预览数据
        """
        try:
            if not self.validate_file(file_path):
                return []
            
            # 临时修改max_rows进行预览
            original_max_rows = SOURCE_CONFIG['max_rows']
            SOURCE_CONFIG['max_rows'] = rows
            
            data_list = self.read_data(file_path)
            
            # 恢复原始配置
            SOURCE_CONFIG['max_rows'] = original_max_rows
            
            return data_list[:rows]
            
        except Exception as e:
            self.logger.error(f"预览数据失败: {e}")
            return []
    
    def get_data_summary(self, file_path: str) -> Dict[str, Any]:
        """
        获取文件数据摘要信息
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            Dict: 数据摘要
        """
        try:
            data_list = self.read_data(file_path)
            
            summary = {
                'total_rows': len(data_list),
                'fields': list(SOURCE_CONFIG['fields'].keys()),
                'preview': data_list[:3] if data_list else [],
                'has_amount': any(row.get('amount') is not None for row in data_list),
            }
            
            # 计算金额总计（如果有金额字段）
            if summary['has_amount']:
                total_amount = 0
                for row in data_list:
                    amount = row.get('amount')
                    if amount is not None:
                        try:
                            # 尝试转换为数字
                            if isinstance(amount, str):
                                # 移除逗号等格式符号
                                amount = amount.replace(',', '').replace('，', '')
                            total_amount += float(amount)
                        except (ValueError, TypeError):
                            pass
                summary['total_amount'] = total_amount
            
            return summary
            
        except Exception as e:
            self.logger.error(f"获取数据摘要失败: {e}")
            return {'total_rows': 0, 'fields': [], 'preview': []} 