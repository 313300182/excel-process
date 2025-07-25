# -*- coding: utf-8 -*-
"""
手工费操作表读取器
从手工费汇总表中提取员工的部位数量、面部数量和考勤数据
"""

import logging
import calendar
from datetime import datetime
from typing import Dict, Any, Optional
import openpyxl
from openpyxl.utils import column_index_from_string


class OperationTableReader:
    """手工费操作表读取器"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
    def read_operation_data(self, file_path: str) -> Dict[str, Any]:
        """
        读取手工费操作表数据（支持两个sheet）
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            Dict[str, Any]: 员工数据 {员工姓名: {部位数量, 面部数量, 考勤数据}}
        """
        workbook = None
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # 获取当前月份天数（用于计算上班天数）
            current_month_days = self._get_current_month_days()
            self.logger.info(f"📅 当前月份天数: {current_month_days} 天")
            
            # 从Sheet1读取手工费数据
            manual_fee_data = self._read_manual_fee_data(workbook)
            
            # 从Sheet2读取考勤数据
            attendance_data = self._read_attendance_data(workbook, current_month_days)
            
            # 合并数据
            combined_data = self._merge_operation_data(manual_fee_data, attendance_data)
            
            # 输出汇总信息
            self.logger.info("=" * 70)
            self.logger.info(f"✅ 操作表读取完成! 共读取 {len(combined_data)} 个员工数据")
            
            if combined_data:
                total_body = sum(data['body_count'] for data in combined_data.values())
                total_face = sum(data['face_count'] for data in combined_data.values())
                total_work_days = sum(data['work_days'] for data in combined_data.values())
                total_rest_days = sum(data['rest_days'] for data in combined_data.values())
                total_actual_absent_days = sum(data['actual_absent_days'] for data in combined_data.values())
                total_training_days = sum(data['training_days'] for data in combined_data.values())
                total_late_count = sum(data['late_count'] for data in combined_data.values())
                
                self.logger.info(f"📊 数据汇总统计:")
                self.logger.info(f"   手工费: 总部位数量={total_body}次, 总面部数量={total_face}次")
                self.logger.info(f"   考勤: 总上班天数={total_work_days}天, 总休息天数={total_rest_days}天, 总实际缺勤天数={total_actual_absent_days}天, 总培训天数={total_training_days}天, 总迟到次数={total_late_count}次")
            
            self.logger.info("=" * 70)
            
            return combined_data
            
        except Exception as e:
            self.logger.error(f"读取手工费操作表失败 {file_path}: {str(e)}")
            raise Exception(f"读取手工费操作表失败: {str(e)}")
        finally:
            # 确保workbook被正确关闭，释放内存
            if workbook:
                try:
                    workbook.close()
                    workbook = None
                except:
                    pass
    
    def _get_current_month_days(self) -> int:
        """
        获取当前月份的天数
        
        Returns:
            int: 当前月份天数
        """
        try:
            now = datetime.now()
            return calendar.monthrange(now.year, now.month)[1]
        except Exception as e:
            self.logger.warning(f"获取当前月份天数失败: {str(e)}")
            return 30  # 默认30天
    
    def _read_manual_fee_data(self, workbook) -> Dict[str, Any]:
        """
        从Sheet1读取手工费数据
        
        Args:
            workbook: Excel工作簿
            
        Returns:
            Dict[str, Any]: 手工费数据
        """
        try:
            # 尝试获取第一个sheet
            if len(workbook.worksheets) == 0:
                raise Exception("工作簿中没有找到工作表")
            
            worksheet = workbook.worksheets[0]  # Sheet1
            sheet_name = worksheet.title
            self.logger.info(f"📋 读取Sheet1手工费数据: {sheet_name}")
            
            manual_fee_data = {}
            
            # 查找表头位置
            header_info = self._find_manual_fee_headers(worksheet)
            if not header_info:
                raise Exception("Sheet1中未找到有效的手工费表头信息")
                
            start_row = header_info['header_row'] + 1
            name_col = header_info['name_col']
            body_count_col = header_info.get('body_count_col')
            face_count_col = header_info.get('face_count_col')
            
            self.logger.info(f"Sheet1表头 - 行: {header_info['header_row']}, 姓名列: {name_col}, 部位数量列: {body_count_col}, 面部数量列: {face_count_col}")
            
            self.logger.info("=" * 50)
            self.logger.info("开始读取Sheet1手工费数据:")
            self.logger.info("=" * 50)
            
            # 读取数据行
            for row_idx in range(start_row, worksheet.max_row + 1):
                try:
                    # 获取员工姓名
                    name_cell = worksheet.cell(row=row_idx, column=name_col)
                    if not name_cell.value:
                        continue
                        
                    employee_name = self._normalize_employee_name(name_cell.value)
                    if not employee_name or employee_name in ['合计', '小计', '总计']:
                        continue
                    
                    # 获取部位数量
                    body_count = 0
                    if body_count_col:
                        body_cell = worksheet.cell(row=row_idx, column=body_count_col)
                        body_count = self._convert_to_number(body_cell.value)
                    
                    # 获取面部数量
                    face_count = 0
                    if face_count_col:
                        face_cell = worksheet.cell(row=row_idx, column=face_count_col)
                        face_count = self._convert_to_number(face_cell.value)
                    
                    # 存储数据
                    manual_fee_data[employee_name] = {
                        'body_count': body_count,
                        'face_count': face_count
                    }
                    
                    # 打印员工数据
                    self.logger.info(f"📋 员工: {employee_name:8} | 部位数量: {body_count:3}次 | 面部数量: {face_count:3}次")
                    
                except Exception as e:
                    self.logger.warning(f"读取Sheet1第 {row_idx} 行数据失败: {str(e)}")
                    continue
            
            self.logger.info(f"✅ Sheet1手工费数据读取完成，共 {len(manual_fee_data)} 个员工")
            return manual_fee_data
            
        except Exception as e:
            self.logger.error(f"读取Sheet1手工费数据失败: {str(e)}")
            return {}
    
    def _read_attendance_data(self, workbook, current_month_days: int) -> Dict[str, Any]:
        """
        从Sheet2读取考勤数据
        
        Args:
            workbook: Excel工作簿
            current_month_days: 当前月份天数
            
        Returns:
            Dict[str, Any]: 考勤数据
        """
        try:
            # 检查是否有第二个sheet
            if len(workbook.worksheets) < 2:
                self.logger.warning("未找到Sheet2，跳过考勤数据读取")
                return {}
            
            worksheet = workbook.worksheets[1]  # Sheet2
            sheet_name = worksheet.title
            self.logger.info(f"📅 读取Sheet2考勤数据: {sheet_name}")
            
            # 获取缺勤计算配置（暂时使用默认值，后续可从配置文件读取）
            base_monthly_rest_days = 4  # 基础月休天数
            current_month_holiday_days = 0  # 当月节日休息天数
            
            self.logger.info(f"📅 缺勤计算参数: 基础月休={base_monthly_rest_days}天, 当月节日={current_month_holiday_days}天")
            
            attendance_data = {}
            
            # 查找表头位置
            header_info = self._find_attendance_headers(worksheet)
            if not header_info:
                self.logger.warning("Sheet2中未找到有效的考勤表头信息")
                return {}
                
            start_row = header_info['header_row'] + 1
            name_col = header_info['name_col']
            rest_days_col = header_info.get('rest_days_col')
            late_count_col = header_info.get('late_count_col') 
            training_days_col = header_info.get('training_days_col')
            personal_tax_col = header_info.get('personal_tax_col')
            
            self.logger.info(f"Sheet2表头 - 行: {header_info['header_row']}, 姓名列: {name_col}, 休息列: {rest_days_col}, 迟到列: {late_count_col}, 培训列: {training_days_col}, 个税列: {personal_tax_col}")
            
            self.logger.info("=" * 50)
            self.logger.info("开始读取Sheet2考勤数据:")
            self.logger.info("=" * 50)
            
            # 读取数据行
            for row_idx in range(start_row, worksheet.max_row + 1):
                try:
                    # 获取员工姓名
                    name_cell = worksheet.cell(row=row_idx, column=name_col)
                    if not name_cell.value:
                        continue
                        
                    employee_name = self._normalize_employee_name(name_cell.value)
                    if not employee_name or employee_name in ['合计', '小计', '总计']:
                        continue
                    
                    # 获取休息天数
                    rest_days = 0
                    if rest_days_col:
                        rest_cell = worksheet.cell(row=row_idx, column=rest_days_col)
                        rest_days = self._convert_to_number(rest_cell.value)
                    
                    # 获取迟到次数
                    late_count = 0
                    if late_count_col:
                        late_cell = worksheet.cell(row=row_idx, column=late_count_col)
                        late_count = self._convert_to_number(late_cell.value)
                    
                    # 获取培训天数
                    training_days = 0
                    if training_days_col:
                        training_cell = worksheet.cell(row=row_idx, column=training_days_col)
                        training_days = self._convert_to_number(training_cell.value)
                    
                    # 获取个人所得税金额
                    personal_tax_amount = 0
                    if personal_tax_col:
                        tax_cell = worksheet.cell(row=row_idx, column=personal_tax_col)
                        personal_tax_amount = self._convert_to_number(tax_cell.value)
                    
                    # 计算实际缺勤天数 = 休息天数 - 基础月休天数 - 当月节日休息天数（可以为负数）
                    actual_absent_days = rest_days - base_monthly_rest_days - current_month_holiday_days
                    
                    # 计算上班天数 = 当前月份天数 - 休息天数
                    work_days = max(0, current_month_days - rest_days)
                    
                    # 存储数据
                    attendance_data[employee_name] = {
                        'rest_days': rest_days,              # 总休息天数（原始数据）
                        'actual_absent_days': actual_absent_days,  # 实际缺勤天数
                        'late_count': late_count,
                        'training_days': training_days,
                        'work_days': work_days,
                        'personal_tax_amount': personal_tax_amount  # 个人所得税金额
                    }
                    
                    # 打印员工数据（详细格式）
                    self.logger.info(f"📅 员工: {employee_name:8} | 上班天数: {work_days:2}天({current_month_days}-{rest_days}) | 休息天数: {rest_days:2}天 | 实际缺勤: {actual_absent_days:2}天({rest_days}-{base_monthly_rest_days}-{current_month_holiday_days}) | 培训天数: {training_days:2}天 | 迟到: {late_count:2}次 | 个税: {personal_tax_amount:.2f}元")
                    
                except Exception as e:
                    self.logger.warning(f"读取Sheet2第 {row_idx} 行数据失败: {str(e)}")
                    continue
            
            self.logger.info(f"✅ Sheet2考勤数据读取完成，共 {len(attendance_data)} 个员工")
            return attendance_data
            
        except Exception as e:
            self.logger.error(f"读取Sheet2考勤数据失败: {str(e)}")
            return {}
    
    def _find_manual_fee_headers(self, worksheet) -> Optional[Dict[str, Any]]:
        """
        查找Sheet1手工费表头信息
        
        Args:
            worksheet: Excel工作表
            
        Returns:
            Optional[Dict[str, Any]]: 表头信息
        """
        try:
            header_info = {
                'header_row': None,
                'name_col': None,
                'body_count_col': None,
                'face_count_col': None
            }
            
            # 在前10行中查找表头
            for row_idx in range(1, min(11, worksheet.max_row + 1)):
                row_data = []
                for col_idx in range(1, min(21, worksheet.max_column + 1)):  # 检查前20列
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    row_data.append(cell.value)
                
                # 检查是否为表头行
                name_found = False
                
                for col_idx, cell_value in enumerate(row_data, 1):
                    if cell_value and isinstance(cell_value, str):
                        cell_str = str(cell_value).strip()
                        
                        # 查找姓名列
                        if any(keyword in cell_str for keyword in ['姓名', '名字', '员工', '操作老师', '老师']):
                            header_info['name_col'] = col_idx
                            name_found = True
                            
                        # 查找部位数量列
                        elif any(keyword in cell_str for keyword in ['部位数量']) and '手工' not in cell_str and '元' not in cell_str:
                            header_info['body_count_col'] = col_idx
                            
                        # 查找面部数量列
                        elif any(keyword in cell_str for keyword in ['面部数量', '面部']):
                            header_info['face_count_col'] = col_idx
                
                # 如果找到了姓名列，认为找到了表头行
                if name_found:
                    header_info['header_row'] = row_idx
                    break
            
            # 验证是否找到了必要的列
            if header_info['header_row'] and header_info['name_col']:
                return header_info
            else:
                return None
                
        except Exception as e:
            self.logger.error(f"查找Sheet1表头失败: {str(e)}")
            return None
    
    def _find_attendance_headers(self, worksheet) -> Optional[Dict[str, Any]]:
        """
        查找Sheet2考勤表头信息
        
        Args:
            worksheet: Excel工作表
            
        Returns:
            Optional[Dict[str, Any]]: 表头信息
        """
        try:
            header_info = {
                'header_row': None,
                'name_col': None,
                'rest_days_col': None,
                'late_count_col': None,
                'training_days_col': None,
                'personal_tax_col': None
            }
            
            self.logger.info("🔍 开始查找Sheet2表头信息...")
            
            # 在前10行中查找表头
            for row_idx in range(1, min(11, worksheet.max_row + 1)):
                row_data = []
                for col_idx in range(1, min(21, worksheet.max_column + 1)):  # 检查前20列
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    row_data.append(cell.value)
                
                # 调试信息：显示当前行内容
                non_empty_cells = [f"列{i+1}:{v}" for i, v in enumerate(row_data) if v is not None]
                if non_empty_cells:
                    self.logger.info(f"🔍 第{row_idx}行内容: {', '.join(non_empty_cells)}")
                
                # 检查是否为表头行
                name_found = False
                
                for col_idx, cell_value in enumerate(row_data, 1):
                    if cell_value and isinstance(cell_value, str):
                        cell_str = str(cell_value).strip()
                        self.logger.debug(f"  检查单元格 列{col_idx}: '{cell_str}'")
                        
                        # 查找姓名列
                        if any(keyword in cell_str for keyword in ['姓名', '名字', '员工', '操作老师', '老师']):
                            header_info['name_col'] = col_idx
                            name_found = True
                            self.logger.info(f"✅ 找到姓名列: 列{col_idx} = '{cell_str}'")
                            
                        # 查找休息天数列
                        elif any(keyword in cell_str for keyword in ['休息', '休假']):
                            header_info['rest_days_col'] = col_idx
                            self.logger.info(f"✅ 找到休息列: 列{col_idx} = '{cell_str}'")
                            
                        # 查找迟到次数列
                        elif any(keyword in cell_str for keyword in ['迟到']):
                            header_info['late_count_col'] = col_idx
                            self.logger.info(f"✅ 找到迟到列: 列{col_idx} = '{cell_str}'")
                            
                        # 查找培训天数列
                        elif any(keyword in cell_str for keyword in ['培训']):
                            header_info['training_days_col'] = col_idx
                            self.logger.info(f"✅ 找到培训列: 列{col_idx} = '{cell_str}'")
                            
                        # 查找个人所得税列
                        elif any(keyword in cell_str for keyword in ['个税', '个人所得税', '所得税', '税金', '税额']):
                            header_info['personal_tax_col'] = col_idx
                            self.logger.info(f"✅ 找到个税列: 列{col_idx} = '{cell_str}'")
                
                # 如果找到了姓名列，认为找到了表头行
                if name_found:
                    header_info['header_row'] = row_idx
                    self.logger.info(f"✅ 确定表头行: 第{row_idx}行")
                    break
            
            # 验证是否找到了必要的列
            if header_info['header_row'] and header_info['name_col']:
                self.logger.info(f"🎯 Sheet2表头识别成功: {header_info}")
                
                # 检查个人所得税列是否找到
                if not header_info['personal_tax_col']:
                    self.logger.warning("⚠️  未找到个人所得税列，将使用默认值0")
                else:
                    self.logger.info(f"✅ 个人所得税列已找到: 列{header_info['personal_tax_col']}")
                
                return header_info
            else:
                self.logger.warning(f"❌ Sheet2表头识别失败: {header_info}")
                return None
                
        except Exception as e:
            self.logger.error(f"查找Sheet2表头失败: {str(e)}")
            return None
    
    def _merge_operation_data(self, manual_fee_data: Dict[str, Any], attendance_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        合并手工费数据和考勤数据
        
        Args:
            manual_fee_data: 手工费数据
            attendance_data: 考勤数据
            
        Returns:
            Dict[str, Any]: 合并后的数据
        """
        combined_data = {}
        
        # 获取所有员工姓名
        all_employees = set(manual_fee_data.keys()) | set(attendance_data.keys())
        
        self.logger.info("=" * 70)
        self.logger.info("📊 合并Sheet1手工费数据和Sheet2考勤数据:")
        self.logger.info("=" * 70)
        
        for employee_name in all_employees:
            # 获取手工费数据
            manual_data = manual_fee_data.get(employee_name, {'body_count': 0, 'face_count': 0})
            
            # 获取考勤数据
            attendance = attendance_data.get(employee_name, {
                'rest_days': 0, 
                'actual_absent_days': 0,
                'late_count': 0, 
                'training_days': 0, 
                'work_days': 0,
                'personal_tax_amount': 0
            })
            
            # 合并数据
            combined_data[employee_name] = {
                'body_count': manual_data['body_count'],
                'face_count': manual_data['face_count'],
                'rest_days': attendance['rest_days'],
                'actual_absent_days': attendance['actual_absent_days'],
                'late_count': attendance['late_count'],
                'training_days': attendance['training_days'],
                'work_days': attendance['work_days'],
                'personal_tax_amount': attendance.get('personal_tax_amount', 0)  # 添加个人所得税
            }
            
            # 打印合并结果
            current_month_days = self._get_current_month_days()
            base_rest = 4  # 基础月休
            holiday_rest = 0  # 节日休息
            personal_tax = attendance.get('personal_tax_amount', 0)
            self.logger.info(f"👤 {employee_name:8} | 手工费[部位:{manual_data['body_count']:2} 面部:{manual_data['face_count']:2}] | 考勤[上班:{attendance['work_days']:2}天({current_month_days}-{attendance['rest_days']}) 休息:{attendance['rest_days']:2}天 实际缺勤:{attendance['actual_absent_days']:2}天({attendance['rest_days']}-{base_rest}-{holiday_rest}) 培训:{attendance['training_days']:2}天 迟到:{attendance['late_count']:2}次] | 个税:{personal_tax:.2f}元")
        
        return combined_data
            
    def _convert_to_number(self, value) -> float:
        """
        将值转换为数字
        
        Args:
            value: 要转换的值
            
        Returns:
            float: 转换后的数字，如果转换失败返回0
        """
        if value is None:
            return 0.0
            
        try:
            if isinstance(value, (int, float)):
                return float(value)
            elif isinstance(value, str):
                # 移除逗号和空格
                cleaned = str(value).replace(',', '').replace(' ', '').strip()
                if cleaned:
                    return float(cleaned)
            return 0.0
        except (ValueError, TypeError):
            return 0.0
    
    def _normalize_employee_name(self, name) -> str:
        """
        标准化员工姓名，去除多余空格和特殊字符
        
        Args:
            name: 原始姓名
            
        Returns:
            str: 标准化后的姓名
        """
        if not name:
            return ""
        
        # 转换为字符串并去除首尾空格
        name_str = str(name).strip()
        
        # 去除中间的所有空格（包括全角空格）
        import re
        name_str = re.sub(r'\s+', '', name_str)
        
        # 去除其他可能的特殊字符
        name_str = name_str.replace('\u3000', '')  # 全角空格
        
        return name_str
            
    def validate_operation_table(self, file_path: str) -> bool:
        """
        验证手工费操作表文件是否有效
        
        Args:
            file_path: 文件路径
            
        Returns:
            bool: 是否有效
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # 检查Sheet1
            if len(workbook.worksheets) == 0:
                return False
            
            worksheet1 = workbook.worksheets[0]
            header_info1 = self._find_manual_fee_headers(worksheet1)
            
            # Sheet1必须有效
            if not header_info1:
                workbook.close()
                return False
            
            # Sheet2是可选的，但如果存在应该有效
            if len(workbook.worksheets) >= 2:
                worksheet2 = workbook.worksheets[1]
                header_info2 = self._find_attendance_headers(worksheet2)
                if not header_info2:
                    self.logger.warning("Sheet2存在但表头无效，将忽略考勤数据")
            
            workbook.close()
            return True
            
        except Exception as e:
            self.logger.error(f"验证手工费操作表失败 {file_path}: {str(e)}")
            return False
            
    def get_employee_operation_data(self, file_path: str, employee_name: str) -> Dict[str, float]:
        """
        获取特定员工的操作数据
        
        Args:
            file_path: 文件路径
            employee_name: 员工姓名
            
        Returns:
            Dict[str, float]: 员工操作数据
        """
        try:
            operation_data = self.read_operation_data(file_path)
            return operation_data.get(employee_name, {
                'body_count': 0, 
                'face_count': 0,
                'rest_days': 0,
                'actual_absent_days': 0,
                'late_count': 0,
                'training_days': 0,
                'work_days': 0
            })
        except Exception as e:
            self.logger.error(f"获取员工 {employee_name} 操作数据失败: {str(e)}")
            return {
                'body_count': 0, 
                'face_count': 0,
                'rest_days': 0,
                'actual_absent_days': 0,
                'late_count': 0,
                'training_days': 0,
                'work_days': 0
            } 